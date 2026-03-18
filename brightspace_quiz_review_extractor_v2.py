#!/usr/bin/env python3
from __future__ import annotations

import argparse
import html
import json
import math
import re
import shutil
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import unquote
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

D2L_NS = "http://desire2learn.com/xsd/d2lcp_v2p0"
HTML_TAG_RE = re.compile(r"<[^>]+>")
WHITESPACE_RE = re.compile(r"\s+")
NORMALIZE_RE = re.compile(r"[^a-z0-9]+")
IMG_SRC_RE = re.compile(
    r"""<img\b[^>]*\bsrc\s*=\s*(?:"([^"]+)"|'([^']+)'|([^\s"'=<>`]+))""",
    re.IGNORECASE,
)
ID_LIKE_TITLE_RE = re.compile(r"^(?:QUES|ITEM|OBJ)_[A-Za-z0-9_:-]+$")
SHEET_ORDER = [
    "quiz_overview",
    "sections_pools",
    "questions",
    "matching_pairs_expanded",
    "pool_members",
    "diagnostics",
    "source_map",
]
DEFAULT_SUGGESTED_ACTION = "Review manually."
TEXT_WIDTH_OVERRIDES = {
    "quiz_overview": {
        "instructions_text": 48,
        "review_notes": 28,
    },
    "sections_pools": {
        "section_title": 28,
        "notes": 36,
    },
    "questions": {
        "question_title_review": 34,
        "stem_text": 54,
        "matching_review_display": 50,
        "ordering_review_display": 50,
        "correct_answer_text": 42,
        "accepted_answers": 42,
        "matching_pairs": 42,
        "ordering_sequence": 32,
        "numeric_tolerance": 32,
        "general_feedback": 42,
        "correct_feedback": 42,
        "incorrect_feedback": 42,
        "answer_specific_feedback": 48,
        "grading_notes": 36,
        "asset_refs": 34,
        "image_refs": 34,
        "image_paths_resolved": 40,
        "image_link_primary": 40,
        "question_payload_json": 60,
        "reviewer_notes": 28,
        "source_hint": 38,
    },
    "matching_pairs_expanded": {
        "question_title_review": 34,
        "prompt": 48,
        "correct_match": 36,
        "image_link_primary": 40,
        "image_refs": 32,
        "image_paths_resolved": 38,
        "source_hint": 38,
    },
    "pool_members": {
        "pool_title": 28,
        "question_title": 28,
        "source_hint": 38,
        "reviewer_notes": 28,
    },
    "diagnostics": {
        "message": 64,
        "suggested_action": 36,
        "source_hint": 38,
    },
    "source_map": {
        "object_title": 32,
        "source_hint": 42,
    },
}


def sanitize_xml(text: str) -> str:
    text = text.lstrip("\ufeff").strip()
    for tag in (
        "questestinterop",
        "objectbank",
        "discussion",
        "grades",
        "rubrics",
        "dropbox",
        "checklists",
        "conditional_release",
        "agents",
        "manifest",
        "cartridge_basiclti_link",
        "quiz",
        "assessment",
    ):
        idx = text.find(f"<{tag}")
        if idx != -1:
            return text[idx:]
    first = text.find("<")
    return text[first:] if first != -1 else text


def parse_xml(path: Path) -> ET.Element:
    text = sanitize_xml(path.read_text(encoding="utf-8-sig", errors="replace"))
    return ET.fromstring(text)


def local_name(tag: str) -> str:
    return tag.split("}")[-1]


def html_to_text(value: Optional[str]) -> str:
    if not value:
        return ""
    text = html.unescape(value)
    text = HTML_TAG_RE.sub(" ", text)
    return WHITESPACE_RE.sub(" ", text).strip()


def normalize_text(value: Optional[str]) -> str:
    return NORMALIZE_RE.sub(" ", html_to_text(value).lower()).strip()


def parse_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def format_number(value: Optional[float]) -> Any:
    if value is None:
        return ""
    if int(value) == value:
        return int(value)
    return round(value, 6)


def unique_preserve(values: Iterable[str]) -> List[str]:
    seen = set()
    out = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return out


def split_semicolon_values(value: Any) -> List[str]:
    return [part.strip() for part in str(value or "").split(";") if part and part.strip()]


def parse_json_field(value: Any, *, default: Any) -> Any:
    if not value:
        return default
    try:
        return json.loads(value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return default


def normalize_whitespace(value: Optional[str]) -> str:
    return WHITESPACE_RE.sub(" ", str(value or "")).strip()


def clean_stem_for_title(value: Optional[str]) -> str:
    text = normalize_whitespace(re.sub(r"\[BLANK_\d+\]", " ", str(value or "")))
    return text.strip(" -:;,")


def is_id_like_title(title: str, question_id: str) -> bool:
    cleaned_title = (title or "").strip()
    if not cleaned_title:
        return True
    if question_id and cleaned_title == question_id:
        return True
    return bool(ID_LIKE_TITLE_RE.fullmatch(cleaned_title))


def derive_question_title_review(question_title: str, question_id: str, stem_text: str) -> str:
    raw_title = normalize_whitespace(question_title)
    if raw_title and not is_id_like_title(raw_title, question_id):
        return raw_title

    stem = clean_stem_for_title(stem_text)
    if not stem:
        return question_id or raw_title

    sentence_match = re.split(r"(?<=[.?!])\s+", stem, maxsplit=1)[0]
    sentence = normalize_whitespace(sentence_match).strip(" -:;")
    if sentence and len(sentence) <= 90:
        return sentence

    words = stem.split()
    shortened = " ".join(words[:12]).strip(" -:;")
    if len(shortened) > 90:
        shortened = shortened[:87].rstrip(" -:;,") + "..."
    return shortened or question_id or raw_title


def qti_metadata(elem: ET.Element) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for field in elem.findall(".//qti_metadatafield"):
        label = field.findtext("fieldlabel", default="").strip()
        entry = field.findtext("fieldentry", default="").strip()
        if label:
            out[label] = entry
    return out


def attr_local(elem: ET.Element, suffix: str, default: str = "") -> str:
    for key, value in elem.attrib.items():
        if key.endswith(suffix):
            return value
    return default


def get_text_from_material(material: Optional[ET.Element]) -> str:
    if material is None:
        return ""
    texts = []
    for mattext in material.findall(".//mattext"):
        cleaned = html_to_text(mattext.text or "")
        if cleaned:
            texts.append(cleaned)
    return " ".join(texts).strip()


def extract_image_refs_from_html(value: Optional[str]) -> List[str]:
    if not value:
        return []
    refs = []
    for match in IMG_SRC_RE.finditer(html.unescape(value)):
        ref = next((group for group in match.groups() if group), "").strip()
        if ref:
            refs.append(ref)
    return unique_preserve(refs)


def collect_image_refs_from_node(node: Optional[ET.Element]) -> List[str]:
    if node is None:
        return []
    refs = []
    for mattext in node.findall(".//mattext"):
        refs.extend(extract_image_refs_from_html(mattext.text or ""))
    for matimg in node.findall(".//matimage"):
        ref = html.unescape((matimg.attrib.get("uri") or matimg.attrib.get("src") or "").strip())
        if ref:
            refs.append(ref)
    return unique_preserve(refs)


def collect_question_image_refs(item: ET.Element) -> List[str]:
    refs = []
    refs.extend(collect_image_refs_from_node(item.find("./presentation")))
    for feedback in item.findall(".//itemfeedback"):
        refs.extend(collect_image_refs_from_node(feedback))
    for answer_key_material in item.findall(".//answer_key_material"):
        refs.extend(collect_image_refs_from_node(answer_key_material))
    return unique_preserve(refs)


def build_export_file_index(export_root: Path) -> Dict[str, Any]:
    relative_lookup: Dict[str, str] = {}
    basename_lookup: Dict[str, List[str]] = defaultdict(list)
    for path in export_root.rglob("*"):
        if not path.is_file():
            continue
        rel_path = path.relative_to(export_root).as_posix()
        relative_lookup[rel_path.lower()] = rel_path
        basename_lookup[path.name.lower()].append(rel_path)
    for matches in basename_lookup.values():
        matches.sort()
    return {
        "relative_lookup": relative_lookup,
        "basename_lookup": dict(basename_lookup),
    }


def normalize_image_ref(raw_ref: str) -> Dict[str, str]:
    raw = html.unescape((raw_ref or "").strip())
    if not raw:
        return {"raw": raw, "error": "empty"}

    candidate = raw.strip().strip("\"'")
    if not candidate:
        return {"raw": raw, "error": "empty"}
    if candidate.lower().startswith("data:"):
        return {"raw": raw, "error": "data_uri"}

    normalized = candidate.replace("\\", "/")
    lower_normalized = normalized.lower()
    if lower_normalized.startswith(("http://", "https://", "file://")):
        marker = "/csfiles/home_dir/"
        marker_index = lower_normalized.find(marker)
        if marker_index == -1:
            return {"raw": raw, "error": "external"}
        normalized = normalized[marker_index + 1 :]
    elif "://" in normalized:
        return {"raw": raw, "error": "external"}

    normalized = normalized.split("#", 1)[0].split("?", 1)[0].strip()
    normalized = unquote(normalized).replace("\\", "/")
    while normalized.startswith("./"):
        normalized = normalized[2:]
    normalized = normalized.lstrip("/")
    if not normalized or normalized in {".", ".."}:
        return {"raw": raw, "error": "empty"}
    return {"raw": raw, "normalized": normalized}


def resolve_image_ref(
    raw_ref: str,
    file_index: Dict[str, Any],
    *,
    question_id: str,
    source_file: str,
    source_hint: str,
) -> Tuple[str, List[Dict[str, str]]]:
    info = normalize_image_ref(raw_ref)
    raw_display = info.get("raw", "")
    if "error" in info:
        error = info["error"]
        if error == "external":
            return "", [
                diagnostic_seed(
                    "unresolved_image_ref",
                    f"Image reference '{raw_display}' points outside the export package and was preserved without resolution.",
                    question_id=question_id,
                    source_file=source_file,
                    source_hint=source_hint,
                    suggested_action="Review the original Brightspace content or preserve the raw image reference for manual follow-up.",
                )
            ]
        return "", [
            diagnostic_seed(
                "malformed_image_ref",
                f"Image reference '{raw_display}' could not be normalized into a local package path.",
                question_id=question_id,
                source_file=source_file,
                source_hint=source_hint,
                suggested_action="Review the raw XML for malformed or unsupported image markup.",
            )
        ]

    normalized = info["normalized"]
    relative_lookup = file_index["relative_lookup"]
    basename_lookup = file_index["basename_lookup"]

    csfiles_candidate = normalized if normalized.lower().startswith("csfiles/home_dir/") else f"csfiles/home_dir/{normalized}"
    resolved = relative_lookup.get(csfiles_candidate.lower())
    if resolved:
        return resolved, []

    export_root_candidate = normalized
    if normalized.lower().startswith("csfiles/home_dir/"):
        export_root_candidate = normalized[len("csfiles/home_dir/") :]
    resolved = relative_lookup.get(export_root_candidate.lower())
    if resolved:
        return resolved, []

    basename = Path(normalized).name
    matches = basename_lookup.get(basename.lower(), [])
    if len(matches) == 1:
        return matches[0], []
    if len(matches) > 1:
        return "", [
            diagnostic_seed(
                "duplicate_image_filename",
                f"Image reference '{raw_display}' matched multiple package files with the same filename: {', '.join(matches)}",
                question_id=question_id,
                source_file=source_file,
                source_hint=source_hint,
                suggested_action="Keep the raw image reference and resolve the intended file manually.",
            ),
            diagnostic_seed(
                "unresolved_image_ref",
                f"Image reference '{raw_display}' was not resolved because multiple package files share the filename '{basename}'.",
                question_id=question_id,
                source_file=source_file,
                source_hint=source_hint,
                suggested_action="Review duplicate filenames in the Brightspace export package manually.",
            ),
        ]

    return "", [
        diagnostic_seed(
            "missing_image_file",
            f"Image reference '{raw_display}' did not resolve to a file in the export package.",
            question_id=question_id,
            source_file=source_file,
            source_hint=source_hint,
            suggested_action="Keep the original export bundle with the review or verify the package contents manually.",
        ),
        diagnostic_seed(
            "unresolved_image_ref",
            f"Image reference '{raw_display}' was preserved but could not be resolved to a local file.",
            question_id=question_id,
            source_file=source_file,
            source_hint=source_hint,
            suggested_action="Review the original Brightspace export and image path manually.",
        ),
    ]


def populate_row_image_fields(
    row: Dict[str, Any],
    *,
    file_index: Dict[str, Any],
    source_file: str,
) -> List[Dict[str, str]]:
    diagnostics: List[Dict[str, str]] = []
    resolved_paths = []
    for raw_ref in split_semicolon_values(row.get("image_refs", "")):
        resolved_path, ref_diags = resolve_image_ref(
            raw_ref,
            file_index,
            question_id=row.get("question_id", ""),
            source_file=source_file,
            source_hint=row.get("source_hint", ""),
        )
        diagnostics.extend(ref_diags)
        if resolved_path:
            resolved_paths.append(resolved_path)

    resolved_paths = unique_preserve(resolved_paths)
    row["image_paths_resolved"] = ";".join(resolved_paths)
    row["image_count"] = len(resolved_paths)
    row["image_link_primary"] = ""
    return diagnostics


def finalize_question_image_links(
    questions: List[Dict[str, Any]],
    *,
    export_root: Path,
    out_dir: Path,
    copy_images_to_assets: bool,
) -> List[Dict[str, str]]:
    diagnostics: List[Dict[str, str]] = []
    copied_targets = set()
    assets_dir = out_dir / "assets"

    for row in questions:
        row["image_link_primary"] = ""
        for rel_path in split_semicolon_values(row.get("image_paths_resolved", "")):
            source_path = export_root / rel_path
            if copy_images_to_assets:
                target_path = assets_dir / Path(rel_path)
                try:
                    if rel_path not in copied_targets:
                        target_path.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(source_path, target_path)
                        copied_targets.add(rel_path)
                    link_target = (Path("assets") / Path(rel_path)).as_posix()
                except OSError as exc:
                    diagnostics.append(
                        fill_diagnostic_context(
                            diagnostic_seed(
                                "image_copy_failed",
                                f"Resolved image '{rel_path}' could not be copied into the output assets folder: {exc}",
                                question_id=row.get("question_id", ""),
                                source_file=row.get("source_quiz_file") or row.get("source_bank_file", ""),
                                source_hint=row.get("source_hint", ""),
                                suggested_action="Review filesystem permissions and package contents, or rerun without image copying.",
                            ),
                            quiz_id=row.get("quiz_id", ""),
                            quiz_title=row.get("quiz_title", ""),
                            section_id=row.get("section_id", ""),
                            question_id=row.get("question_id", ""),
                        )
                    )
                    continue
            else:
                link_target = str(source_path.resolve())

            if not row["image_link_primary"]:
                row["image_link_primary"] = link_target

    return diagnostics


def format_matching_review_display(row: Dict[str, Any]) -> str:
    pairs = parse_json_field(row.get("matching_pairs", ""), default=[])
    if not isinstance(pairs, list):
        return ""

    lines = []
    for pair in pairs:
        prompt = normalize_whitespace(pair.get("prompt", ""))
        correct_values = [normalize_whitespace(value) for value in pair.get("correct", []) if normalize_whitespace(value)]
        correct_text = "; ".join(correct_values) if correct_values else "(unresolved)"
        if prompt:
            lines.append(f"{prompt} -> {correct_text}")
    return "\n".join(lines)


def choice_text_from_key(row: Dict[str, Any], choices: Dict[str, Dict[str, Any]], key: str) -> str:
    choice = choices.get(key, {}) if isinstance(choices, dict) else {}
    text = normalize_whitespace(choice.get("text", ""))
    if text:
        return text
    fallback_key = key.lower()
    return normalize_whitespace(row.get(f"choice_{fallback_key}", ""))


def derive_ordering_sequence_from_payload(payload: Dict[str, Any], ident_to_key: Dict[str, str]) -> List[str]:
    position_to_key: Dict[int, str] = {}
    sequence = []
    for answer in payload.get("conditions", []):
        varname = answer.get("varname", "")
        score = parse_number(answer.get("score"))
        is_correct = (varname == "D2L_Correct") or (not varname and score is not None and score > 0)
        if not is_correct:
            continue
        respident = answer.get("respident", "")
        if respident in ident_to_key:
            position = parse_number(answer.get("value"))
            if position is not None and int(position) == position and position >= 1:
                position_to_key[int(position)] = ident_to_key[respident]
                continue
        value = answer.get("value", "")
        if value in ident_to_key:
            sequence.append(ident_to_key[value])
    if position_to_key:
        return [position_to_key[position] for position in sorted(position_to_key)]
    return unique_preserve(sequence)


def format_ordering_review_display(row: Dict[str, Any]) -> str:
    payload = parse_json_field(row.get("question_payload_json", ""), default={})
    if not isinstance(payload, dict):
        payload = {}

    choices = payload.get("choices", {})
    if not isinstance(choices, dict):
        choices = {}
    ident_to_key = {
        choice.get("ident", ""): key
        for key, choice in choices.items()
        if isinstance(choice, dict) and choice.get("ident")
    }

    sequence = split_semicolon_values(row.get("ordering_sequence", ""))
    if not sequence:
        payload_sequence = payload.get("sequence", [])
        if isinstance(payload_sequence, list):
            sequence = [str(value) for value in payload_sequence if str(value).strip()]
    if not sequence and ident_to_key:
        sequence = derive_ordering_sequence_from_payload(payload, ident_to_key)

    best_effort = bool(payload.get("best_effort"))
    if not sequence:
        if payload.get("conditions"):
            return "Best-effort only: Brightspace preserved ordering conditions, but not a reviewer-friendly ordered item list. See question_payload_json."
        return ""

    lines = []
    if best_effort:
        lines.append("Best-effort sequence (presentation order):")
    for index, key in enumerate(sequence, start=1):
        label = choice_text_from_key(row, choices, key) or f"Item {index}"
        lines.append(f"{index}. {label}")
    return "\n".join(lines)


def enrich_question_rows_for_review(questions: List[Dict[str, Any]]) -> None:
    for row in questions:
        row["question_title_review"] = derive_question_title_review(
            str(row.get("question_title", "")),
            str(row.get("question_id", "")),
            str(row.get("stem_text", "")),
        )
        row["has_image"] = "yes" if int(row.get("image_count") or 0) > 0 else "no"
        row["matching_review_display"] = (
            format_matching_review_display(row) if row.get("question_type") == "Matching" else ""
        )
        row["ordering_review_display"] = (
            format_ordering_review_display(row) if row.get("question_type") == "Ordering" else ""
        )


def build_matching_pairs_expanded_rows(questions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for question in questions:
        pairs = parse_json_field(question.get("matching_pairs", ""), default=[])
        if question.get("question_type") != "Matching" or not isinstance(pairs, list):
            continue
        for prompt_order, pair in enumerate(pairs, start=1):
            prompt = normalize_whitespace(pair.get("prompt", ""))
            correct_values = [normalize_whitespace(value) for value in pair.get("correct", []) if normalize_whitespace(value)]
            if not correct_values:
                correct_values = ["(unresolved)"]
            for match_order, correct_match in enumerate(correct_values, start=1):
                rows.append(
                    {
                        "quiz_id": question.get("quiz_id", ""),
                        "quiz_title": question.get("quiz_title", ""),
                        "section_id": question.get("section_id", ""),
                        "section_title": question.get("section_title", ""),
                        "question_order": question.get("question_order", ""),
                        "question_id": question.get("question_id", ""),
                        "question_title_review": question.get("question_title_review", ""),
                        "question_title": question.get("question_title", ""),
                        "prompt_order": prompt_order,
                        "match_order": match_order,
                        "prompt": prompt,
                        "correct_match": correct_match,
                        "has_image": question.get("has_image", "no"),
                        "image_count": question.get("image_count", 0),
                        "image_refs": question.get("image_refs", ""),
                        "image_paths_resolved": question.get("image_paths_resolved", ""),
                        "image_link_primary": question.get("image_link_primary", ""),
                        "source_location": question.get("source_location", ""),
                        "source_quiz_file": question.get("source_quiz_file", ""),
                        "source_bank_file": question.get("source_bank_file", ""),
                        "source_hint": question.get("source_hint", ""),
                    }
                )
    rows.sort(
        key=lambda row: (
            str(row.get("quiz_id", "")),
            str(row.get("section_id", "")),
            int(parse_number(row.get("question_order")) or 0),
            int(parse_number(row.get("prompt_order")) or 0),
            int(parse_number(row.get("match_order")) or 0),
        )
    )
    return rows


def diagnostic_seed(
    issue_type: str,
    message: str,
    *,
    severity: str = "warning",
    question_id: str = "",
    source_file: str = "",
    source_hint: str = "",
    suggested_action: str = DEFAULT_SUGGESTED_ACTION,
) -> Dict[str, str]:
    return {
        "severity": severity,
        "quiz_id": "",
        "quiz_title": "",
        "section_id": "",
        "question_id": question_id,
        "issue_type": issue_type,
        "message": message,
        "source_file": source_file,
        "source_hint": source_hint,
        "suggested_action": suggested_action,
    }


def fill_diagnostic_context(
    diag: Dict[str, str],
    *,
    quiz_id: str,
    quiz_title: str,
    section_id: str = "",
    question_id: str = "",
) -> Dict[str, str]:
    out = dict(diag)
    out["quiz_id"] = quiz_id
    out["quiz_title"] = quiz_title
    if section_id:
        out["section_id"] = section_id
    if question_id:
        out["question_id"] = question_id
    return out


def make_source_map_row(
    *,
    object_type: str,
    object_id: str,
    object_title: str,
    source_file: str,
    source_hint: str,
    resolved_to_sheet: str,
    resolved_to_key: str,
    quiz_id: str = "",
    quiz_title: str = "",
) -> Dict[str, str]:
    return {
        "object_type": object_type,
        "object_id": object_id,
        "object_title": object_title,
        "quiz_id": quiz_id,
        "quiz_title": quiz_title,
        "source_file": source_file,
        "source_hint": source_hint,
        "resolved_to_sheet": resolved_to_sheet,
        "resolved_to_key": resolved_to_key,
    }


def get_presentation_parts(item: ET.Element) -> Tuple[str, List[Dict[str, str]], List[str]]:
    presentation = item.find("./presentation")
    if presentation is None:
        return "", [], []

    blanks: List[Dict[str, str]] = []
    assets = set()

    def walk(node: ET.Element, blank_counter: List[int]) -> List[str]:
        parts: List[str] = []
        for child in list(node):
            tag = local_name(child.tag)
            if tag == "material":
                text = get_text_from_material(child)
                if text:
                    parts.append(text)
                for matimg in child.findall(".//matimage"):
                    uri = matimg.attrib.get("uri") or matimg.attrib.get("src") or ""
                    if uri:
                        assets.add(uri)
            elif tag in ("response_str", "response_num"):
                blank_counter[0] += 1
                render_fib = child.find(".//render_fib")
                response_label = child.find(".//response_label")
                blanks.append(
                    {
                        "blank_index": blank_counter[0],
                        "resp_ident": child.attrib.get("ident", ""),
                        "response_label_ident": response_label.attrib.get("ident", "") if response_label is not None else "",
                        "fibtype": render_fib.attrib.get("fibtype", "") if render_fib is not None else "",
                        "response_type": tag,
                    }
                )
                parts.append(f"[BLANK_{blank_counter[0]}]")
            elif tag == "response_grp":
                continue
            elif tag == "response_lid":
                continue
            else:
                parts.extend(walk(child, blank_counter))
        return parts

    parts = walk(presentation, [0])
    stem = WHITESPACE_RE.sub(" ", " ".join(p for p in parts if p)).strip()
    return stem, blanks, sorted(assets)


def get_itemfeedback_map(item: ET.Element) -> Dict[str, str]:
    feedbacks: Dict[str, str] = {}
    for feedback in item.findall(".//itemfeedback"):
        ident = feedback.attrib.get("ident", "")
        texts = []
        for mattext in feedback.findall(".//mattext"):
            cleaned = html_to_text(mattext.text or "")
            if cleaned:
                texts.append(cleaned)
        if ident and texts:
            feedbacks[ident] = " | ".join(texts)
    return feedbacks


def classify_feedbacks(item: ET.Element, source_file: str, source_hint: str, question_id: str) -> Tuple[str, str, str, str, List[Dict[str, str]]]:
    feedback_map = get_itemfeedback_map(item)
    if not feedback_map:
        return "", "", "", "", []

    general_ids = set()
    correct_ids = set()
    incorrect_ids = set()
    answer_specific_parts: List[str] = []
    referenced_ids = set()
    diagnostics: List[Dict[str, str]] = []

    for condition in item.findall("./resprocessing/respcondition"):
        setvars = condition.findall("./setvar")
        numeric_scores = []
        for setvar in setvars:
            score = parse_number(setvar.text)
            if score is not None:
                numeric_scores.append(score)
        varnames = {setvar.attrib.get("varname", "") for setvar in setvars}
        chosen_values = unique_preserve((varequal.text or "").strip() for varequal in condition.findall(".//varequal"))
        refs = [display.attrib.get("linkrefid", "") for display in condition.findall("./displayfeedback")]
        refs = [ref for ref in refs if ref in feedback_map]
        if not refs:
            continue
        referenced_ids.update(refs)

        if "D2L_Correct" in varnames or any(score > 0 for score in numeric_scores):
            correct_ids.update(refs)
        elif "D2L_Incorrect" in varnames or (numeric_scores and all(score <= 0 for score in numeric_scores)):
            incorrect_ids.update(refs)
        elif chosen_values:
            label = ",".join(chosen_values)
            for ref in refs:
                answer_specific_parts.append(f"{label}: {feedback_map[ref]}")
        else:
            general_ids.update(refs)

    unused_ids = set(feedback_map) - referenced_ids
    if unused_ids:
        if correct_ids or incorrect_ids or answer_specific_parts:
            general_ids.update(unused_ids)
            diagnostics.append(
                diagnostic_seed(
                    "feedback_partial_collapse",
                    "Some feedback blocks were not linked to a specific outcome and were preserved as general feedback.",
                    question_id=question_id,
                    source_file=source_file,
                    source_hint=source_hint,
                    suggested_action="Verify whether Brightspace exported separate feedback channels for this item.",
                )
            )
        else:
            general_ids.update(unused_ids)

    overlapping = (correct_ids & incorrect_ids) | (correct_ids & general_ids) | (incorrect_ids & general_ids)
    if overlapping:
        diagnostics.append(
            diagnostic_seed(
                "feedback_ambiguous_classification",
                "One or more feedback blocks were referenced by multiple feedback channels.",
                question_id=question_id,
                source_file=source_file,
                source_hint=source_hint,
                suggested_action="Review feedback routing manually.",
            )
        )

    general_feedback = " | ".join(feedback_map[ident] for ident in sorted(general_ids))
    correct_feedback = " | ".join(feedback_map[ident] for ident in sorted(correct_ids))
    incorrect_feedback = " | ".join(feedback_map[ident] for ident in sorted(incorrect_ids))
    answer_specific_feedback = " | ".join(answer_specific_parts)
    return general_feedback, correct_feedback, incorrect_feedback, answer_specific_feedback, diagnostics


def get_choice_text(response_label: ET.Element) -> str:
    return get_text_from_material(response_label.find("./flow_mat/material")) or get_text_from_material(response_label.find("./material"))


def extract_choice_sets(item: ET.Element) -> Tuple[Dict[str, Dict[str, str]], Dict[str, str], str]:
    response_lid = item.find(".//response_lid")
    rcardinality = response_lid.attrib.get("rcardinality", "") if response_lid is not None else ""
    choices: Dict[str, Dict[str, str]] = {}
    ident_to_key: Dict[str, str] = {}
    index = 0
    for response_label in item.findall(".//response_lid//response_label"):
        index += 1
        key = chr(ord("A") + index - 1) if index <= 26 else f"OPT_{index}"
        ident = response_label.attrib.get("ident", "")
        text = get_choice_text(response_label)
        choices[key] = {"ident": ident, "text": text}
        if ident:
            ident_to_key[ident] = key
    return choices, ident_to_key, rcardinality


def extract_ordering_choices(item: ET.Element) -> Tuple[Dict[str, Dict[str, str]], Dict[str, str]]:
    response_labels = item.findall(".//response_lid//response_label")
    if not response_labels:
        ordered_groups = [
            group
            for group in item.findall(".//response_grp")
            if group.attrib.get("rcardinality", "").lower() == "ordered"
        ]
        if ordered_groups:
            response_labels = ordered_groups[0].findall(".//response_label")

    choices: Dict[str, Dict[str, str]] = {}
    ident_to_key: Dict[str, str] = {}
    for index, response_label in enumerate(response_labels, start=1):
        key = chr(ord("A") + index - 1) if index <= 26 else f"OPT_{index}"
        ident = response_label.attrib.get("ident", "")
        choices[key] = {"ident": ident, "text": get_choice_text(response_label)}
        if ident:
            ident_to_key[ident] = key
    return choices, ident_to_key


def gather_respcondition_answers(item: ET.Element) -> List[Dict[str, Any]]:
    answers: List[Dict[str, Any]] = []
    for condition_index, condition in enumerate(item.findall("./resprocessing/respcondition"), start=1):
        setvars = condition.findall("./setvar")
        score_texts = [" ".join((setvar.text or "").split()) for setvar in setvars if (setvar.text or "").strip()]
        score_text = ";".join(score_texts)
        refs = [display.attrib.get("linkrefid", "") for display in condition.findall("./displayfeedback")]

        for varequal in condition.findall(".//varequal"):
            answers.append(
                {
                    "condition_index": condition_index,
                    "respident": varequal.attrib.get("respident", ""),
                    "value": (varequal.text or "").strip(),
                    "case": varequal.attrib.get("case", ""),
                    "score": score_text,
                    "varname": setvars[0].attrib.get("varname", "") if setvars else "",
                    "action": setvars[0].attrib.get("action", "") if setvars else "",
                    "displayfeedback_refs": refs,
                }
            )

        for operator in ("vargte", "vargt", "varlte", "varlt"):
            for comparator in condition.findall(f".//{operator}"):
                answers.append(
                    {
                        "condition_index": condition_index,
                        "respident": comparator.attrib.get("respident", ""),
                        "value": (comparator.text or "").strip(),
                        "case": "",
                        "score": score_text,
                        "varname": setvars[0].attrib.get("varname", "") if setvars else "",
                        "action": setvars[0].attrib.get("action", "") if setvars else "",
                        "operator": operator,
                        "displayfeedback_refs": refs,
                    }
                )
    return answers


def add_common_feedback(
    item: ET.Element,
    base: Dict[str, Any],
    diagnostics: List[Dict[str, str]],
) -> None:
    general, correct, incorrect, answer_specific, feedback_diags = classify_feedbacks(
        item,
        base["source_quiz_file"] or base["source_bank_file"],
        base["source_hint"],
        base["question_id"],
    )
    base["general_feedback"] = general
    base["correct_feedback"] = correct
    base["incorrect_feedback"] = incorrect
    base["answer_specific_feedback"] = answer_specific
    diagnostics.extend(feedback_diags)


def parse_multiple_choice(item: ET.Element, base: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    diagnostics: List[Dict[str, str]] = []
    choices, ident_to_key, rcardinality = extract_choice_sets(item)
    answers = gather_respcondition_answers(item)
    correct_keys: List[str] = []
    answer_specific_parts: List[str] = []
    feedback_map = get_itemfeedback_map(item)

    for answer in answers:
        score = parse_number(answer.get("score"))
        ident = answer.get("value", "")
        if ident in ident_to_key and score is not None and score > 0:
            correct_keys.append(ident_to_key[ident])
        for ref in answer.get("displayfeedback_refs", []):
            if ref in feedback_map and ident:
                choice_key = ident_to_key.get(ident, ident)
                answer_specific_parts.append(f"{choice_key}: {feedback_map[ref]}")

    correct_keys = unique_preserve(correct_keys)
    if not correct_keys and choices:
        diagnostics.append(
            diagnostic_seed(
                "choice_answer_unresolved",
                "No positively-scored choice condition was found for a choice-based question.",
                question_id=base["question_id"],
                source_file=base["source_quiz_file"] or base["source_bank_file"],
                source_hint=base["source_hint"],
                suggested_action="Verify scoring conditions manually.",
            )
        )

    base.update(
        {
            "choice_a": choices.get("A", {}).get("text", ""),
            "choice_b": choices.get("B", {}).get("text", ""),
            "choice_c": choices.get("C", {}).get("text", ""),
            "choice_d": choices.get("D", {}).get("text", ""),
            "choice_e": choices.get("E", {}).get("text", ""),
            "choice_f": choices.get("F", {}).get("text", ""),
            "correct_answer_key": ";".join(correct_keys),
            "all_correct_keys": ";".join(correct_keys),
            "correct_answer_text": " | ".join(choices[key]["text"] for key in correct_keys if key in choices),
            "response_schema": "multi_select" if rcardinality.lower() == "multiple" or base["question_type"] in ("Multiple Response", "Multi-Select", "Multi Select") else "single_select",
            "question_payload_json": json.dumps(
                {
                    "choices": choices,
                    "scoring_conditions": answers,
                    "rcardinality": rcardinality,
                },
                ensure_ascii=False,
            ),
        }
    )
    add_common_feedback(item, base, diagnostics)
    if answer_specific_parts:
        base["answer_specific_feedback"] = " | ".join(
            part for part in [base["answer_specific_feedback"], " | ".join(answer_specific_parts)] if part
        )
    return base, diagnostics


def parse_short_or_fitb(item: ET.Element, base: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    diagnostics: List[Dict[str, str]] = []
    stem, blanks, assets = get_presentation_parts(item)
    accepted_by_resp: Dict[str, List[str]] = defaultdict(list)
    constraints_by_resp: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    case_by_resp: Dict[str, str] = {}

    for answer in gather_respcondition_answers(item):
        respident = answer.get("respident", "")
        if answer.get("operator"):
            constraints_by_resp[respident].append(
                {"operator": answer["operator"], "value": answer["value"]}
            )
        elif answer.get("value"):
            accepted_by_resp[respident].append(answer["value"])
            if answer.get("case"):
                case_by_resp[respident] = answer["case"]

    accepted_answers = []
    for blank in blanks:
        respident = blank["response_label_ident"] or blank["resp_ident"]
        accepted_values = unique_preserve(
            accepted_by_resp.get(respident, []) or accepted_by_resp.get(blank["resp_ident"], [])
        )
        numeric_constraints = constraints_by_resp.get(respident, []) or constraints_by_resp.get(blank["resp_ident"], [])
        accepted_answers.append(
            {
                "blank_index": blank["blank_index"],
                "resp_ident": respident or blank["resp_ident"],
                "accepted_values": accepted_values,
                "numeric_constraints": numeric_constraints,
                "fibtype": blank["fibtype"],
                "response_type": blank["response_type"],
                "case_sensitive": case_by_resp.get(respident, case_by_resp.get(blank["resp_ident"], "")),
            }
        )

    if not accepted_answers:
        diagnostics.append(
            diagnostic_seed(
                "text_response_structure_unresolved",
                "No blank/response structure could be resolved for this text-response item.",
                question_id=base["question_id"],
                source_file=base["source_quiz_file"] or base["source_bank_file"],
                source_hint=base["source_hint"],
                suggested_action="Review the raw question payload for response structure details.",
            )
        )

    unresolved_blanks = [
        answer
        for answer in accepted_answers
        if not answer["accepted_values"] and not answer["numeric_constraints"]
    ]
    if unresolved_blanks:
        diagnostics.append(
            diagnostic_seed(
                "accepted_answer_unresolved",
                "One or more blanks did not expose accepted answers or numeric constraints.",
                question_id=base["question_id"],
                source_file=base["source_quiz_file"] or base["source_bank_file"],
                source_hint=base["source_hint"],
                suggested_action="Review the text-response scoring conditions manually.",
            )
        )

    accepted_text_parts = []
    for answer in accepted_answers:
        if answer["accepted_values"]:
            answer_text = ", ".join(answer["accepted_values"])
        elif answer["numeric_constraints"]:
            answer_text = json.dumps(answer["numeric_constraints"], ensure_ascii=False)
        else:
            answer_text = "(unresolved)"
        accepted_text_parts.append(f"BLANK_{answer['blank_index']}: {answer_text}")

    base.update(
        {
            "stem_text": stem or base["stem_text"],
            "correct_answer_text": " | ".join(accepted_text_parts),
            "accepted_answers": json.dumps(accepted_answers, ensure_ascii=False),
            "response_schema": "fill_in_blank" if len(accepted_answers) != 1 or base["question_type"] == "Fill in the Blanks" else "short_text",
            "question_payload_json": json.dumps(
                {
                    "blanks": accepted_answers,
                    "assets": assets,
                    "conditions": gather_respcondition_answers(item),
                },
                ensure_ascii=False,
            ),
        }
    )
    add_common_feedback(item, base, diagnostics)
    return base, diagnostics


def parse_matching(item: ET.Element, base: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    diagnostics: List[Dict[str, str]] = []
    presentation = item.find("./presentation/flow")
    first_material = presentation.find("./material") if presentation is not None else None
    if first_material is not None:
        base["stem_text"] = get_text_from_material(first_material) or base["stem_text"]

    groups = []
    option_lookup: Dict[str, str] = {}
    for response_group in item.findall(".//response_grp"):
        prompt = get_text_from_material(response_group.find("./material"))
        respident = response_group.attrib.get("respident", "")
        options = []
        for response_label in response_group.findall(".//response_label"):
            ident = response_label.attrib.get("ident", "")
            text = get_text_from_material(response_label.find("./flow_mat/material")) or get_text_from_material(response_label.find("./material"))
            option_lookup[ident] = text
            options.append({"ident": ident, "text": text})
        groups.append({"prompt": prompt, "respident": respident, "options": options})

    correct_by_prompt: Dict[str, List[str]] = defaultdict(list)
    for answer in gather_respcondition_answers(item):
        varname = answer.get("varname", "")
        score = parse_number(answer.get("score"))
        if varname:
            if varname != "D2L_Correct":
                continue
        elif score is None or score <= 0:
            continue
        value = answer.get("value", "")
        respident = answer.get("respident", "")
        if respident and value:
            correct_by_prompt[respident].append(value)

    pairs = []
    for group in groups:
        correct_idents = unique_preserve(correct_by_prompt.get(group["respident"], []))
        correct_texts = [option_lookup.get(ident, ident) for ident in correct_idents]
        if not correct_texts:
            diagnostics.append(
                diagnostic_seed(
                    "matching_pair_unresolved",
                    "A matching prompt did not expose a resolvable correct option.",
                    question_id=base["question_id"],
                    source_file=base["source_quiz_file"] or base["source_bank_file"],
                    source_hint=base["source_hint"],
                    suggested_action="Review the matching scoring logic manually.",
                )
            )
        pairs.append(
            {
                "prompt": group["prompt"],
                "correct": correct_texts,
                "respident": group["respident"],
            }
        )

    base.update(
        {
            "matching_pairs": json.dumps(pairs, ensure_ascii=False),
            "correct_answer_text": " | ".join(
                f"{pair['prompt']} => {', '.join(pair['correct']) if pair['correct'] else '(unresolved)'}"
                for pair in pairs
            ),
            "response_schema": "matching",
            "question_payload_json": json.dumps({"groups": groups}, ensure_ascii=False),
        }
    )
    add_common_feedback(item, base, diagnostics)
    return base, diagnostics


def parse_long_answer(item: ET.Element, base: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    diagnostics: List[Dict[str, str]] = []
    response_extension = item.find("./presentation/flow/response_extension")
    notes = []
    if response_extension is not None:
        for child in list(response_extension):
            tag = local_name(child.tag)
            text = (child.text or "").strip()
            if text:
                notes.append(f"{tag}={text}")
    base.update(
        {
            "grading_notes": "; ".join(notes),
            "response_schema": "essay",
            "question_payload_json": json.dumps({"manual_grading": True, "response_extension_notes": notes}, ensure_ascii=False),
        }
    )
    add_common_feedback(item, base, diagnostics)
    return base, diagnostics


def parse_numeric(item: ET.Element, base: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    diagnostics: List[Dict[str, str]] = []
    accepted_values = []
    constraints = []
    for answer in gather_respcondition_answers(item):
        if answer.get("operator"):
            constraints.append(
                {
                    "operator": answer["operator"],
                    "value": answer["value"],
                    "respident": answer["respident"],
                }
            )
        elif answer.get("value"):
            accepted_values.append(answer["value"])

    accepted_values = unique_preserve(accepted_values)
    if not accepted_values and not constraints:
        diagnostics.append(
            diagnostic_seed(
                "numeric_answer_unresolved",
                "Numeric/arithmetic scoring conditions were not detected.",
                question_id=base["question_id"],
                source_file=base["source_quiz_file"] or base["source_bank_file"],
                source_hint=base["source_hint"],
                suggested_action="Review the raw numeric response conditions manually.",
            )
        )
    if constraints and not accepted_values:
        diagnostics.append(
            diagnostic_seed(
                "numeric_best_effort",
                "Only numeric comparison/tolerance constraints were preserved for this item.",
                question_id=base["question_id"],
                source_file=base["source_quiz_file"] or base["source_bank_file"],
                source_hint=base["source_hint"],
                suggested_action="Review tolerance logic in the raw payload if exact numeric semantics matter.",
            )
        )

    base.update(
        {
            "numeric_answer": ";".join(accepted_values),
            "numeric_tolerance": json.dumps(constraints, ensure_ascii=False) if constraints else "",
            "correct_answer_text": ";".join(accepted_values) or json.dumps(constraints, ensure_ascii=False),
            "response_schema": "numeric",
            "question_payload_json": json.dumps(
                {
                    "accepted_values": accepted_values,
                    "constraints": constraints,
                    "conditions": gather_respcondition_answers(item),
                },
                ensure_ascii=False,
            ),
        }
    )
    add_common_feedback(item, base, diagnostics)
    return base, diagnostics


def parse_ordering(item: ET.Element, base: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    diagnostics: List[Dict[str, str]] = []
    choices, ident_to_key = extract_ordering_choices(item)
    sequence = []
    sequence_by_position: Dict[int, str] = {}
    for answer in gather_respcondition_answers(item):
        varname = answer.get("varname", "")
        score = parse_number(answer.get("score"))
        is_correct = (varname == "D2L_Correct") or (not varname and score is not None and score > 0)
        if not is_correct:
            continue
        if answer.get("respident") in ident_to_key:
            position = parse_number(answer.get("value"))
            if position is not None and int(position) == position and position >= 1:
                sequence_by_position[int(position)] = ident_to_key[answer["respident"]]
                continue
        if answer.get("value") in ident_to_key:
            sequence.append(ident_to_key[answer["value"]])
    if sequence_by_position:
        sequence = [sequence_by_position[position] for position in sorted(sequence_by_position)]
    sequence = unique_preserve(sequence)
    best_effort = False
    if not sequence:
        sequence = list(choices.keys())
        best_effort = True
        diagnostics.append(
            diagnostic_seed(
                "ordering_best_effort",
                "Ordering question lacked an explicit positively-scored sequence; presentation order was preserved as a best-effort fallback.",
                question_id=base["question_id"],
                source_file=base["source_quiz_file"] or base["source_bank_file"],
                source_hint=base["source_hint"],
                suggested_action="Review the raw ordering conditions manually.",
            )
        )

    base.update(
        {
            "choice_a": choices.get("A", {}).get("text", ""),
            "choice_b": choices.get("B", {}).get("text", ""),
            "choice_c": choices.get("C", {}).get("text", ""),
            "choice_d": choices.get("D", {}).get("text", ""),
            "choice_e": choices.get("E", {}).get("text", ""),
            "choice_f": choices.get("F", {}).get("text", ""),
            "ordering_sequence": ";".join(sequence),
            "correct_answer_key": ";".join(sequence),
            "correct_answer_text": " | ".join(choices[key]["text"] for key in sequence if key in choices),
            "response_schema": "ordering",
            "question_payload_json": json.dumps(
                {
                    "choices": choices,
                    "sequence": sequence,
                    "best_effort": best_effort,
                    "conditions": gather_respcondition_answers(item),
                },
                ensure_ascii=False,
            ),
        }
    )
    add_common_feedback(item, base, diagnostics)
    return base, diagnostics


def build_base_row(
    item: ET.Element,
    source_location: str,
    source_quiz_file: str,
    source_bank_file: str = "",
    source_hint: str = "",
) -> Dict[str, Any]:
    metadata = qti_metadata(item)
    question_type = metadata.get("qmd_questiontype", "") or "Unknown"
    question_id = item.attrib.get("label") or item.attrib.get("ident") or attr_local(item, "id", "")
    question_title = item.attrib.get("title") or item.attrib.get("label") or item.attrib.get("ident") or ""
    stem_text, _, assets = get_presentation_parts(item)
    image_refs = collect_question_image_refs(item)
    if not stem_text:
        stem_text = html_to_text(item.findtext("./presentation/flow/material/mattext", default=""))
    return {
        "quiz_id": "",
        "quiz_title": "",
        "section_id": "",
        "section_title": "",
        "question_order": "",
        "question_id": question_id,
        "question_title": question_title,
        "question_type": question_type,
        "source_location": source_location,
        "source_hint": source_hint,
        "source_quiz_file": source_quiz_file,
        "source_bank_file": source_bank_file,
        "points": metadata.get("qmd_weighting", ""),
        "stem_text": stem_text,
        "choice_a": "",
        "choice_b": "",
        "choice_c": "",
        "choice_d": "",
        "choice_e": "",
        "choice_f": "",
        "correct_answer_key": "",
        "correct_answer_text": "",
        "all_correct_keys": "",
        "accepted_answers": "",
        "matching_pairs": "",
        "ordering_sequence": "",
        "numeric_answer": "",
        "numeric_tolerance": "",
        "response_schema": "",
        "general_feedback": "",
        "correct_feedback": "",
        "incorrect_feedback": "",
        "answer_specific_feedback": "",
        "grading_notes": "",
        "metadata_difficulty": item.findtext(f"./itemproc_extension/{{{D2L_NS}}}difficulty", default="") or metadata.get("qmd_difficulty", ""),
        "metadata_mandatory": item.findtext(f"./itemproc_extension/{{{D2L_NS}}}ismandatory", default=""),
        "asset_refs": ";".join(assets),
        "image_refs": ";".join(image_refs),
        "image_paths_resolved": "",
        "image_count": 0,
        "image_link_primary": "",
        "question_title_review": "",
        "matching_review_display": "",
        "ordering_review_display": "",
        "has_image": "no",
        "question_payload_json": "",
        "review_status": "pending",
        "reviewer_notes": "",
    }


def parse_item(
    item: ET.Element,
    source_location: str,
    source_quiz_file: str,
    source_bank_file: str = "",
    source_hint: str = "",
) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    base = build_base_row(item, source_location, source_quiz_file, source_bank_file, source_hint)
    question_type = base["question_type"].lower()

    if question_type in ("multiple choice", "true/false", "multi-select", "multi select", "multiple response"):
        return parse_multiple_choice(item, base)
    if question_type in ("short answer", "fill in the blanks", "fill in the blank"):
        return parse_short_or_fitb(item, base)
    if question_type == "long answer":
        return parse_long_answer(item, base)
    if question_type == "matching":
        return parse_matching(item, base)
    if question_type in ("ordering", "order", "sequence"):
        return parse_ordering(item, base)
    if question_type in ("numeric", "arithmetic", "numeric/arithmetic"):
        return parse_numeric(item, base)

    answers = gather_respcondition_answers(item)
    diagnostics: List[Dict[str, str]] = [
        diagnostic_seed(
            "question_type_best_effort",
            f"Question type '{base['question_type']}' was parsed using best-effort inference.",
            question_id=base["question_id"],
            source_file=source_quiz_file or source_bank_file,
            source_hint=source_hint,
            suggested_action="Review the question payload JSON if exact semantics matter.",
        )
    ]
    if any(answer.get("operator") for answer in answers):
        row, parse_diags = parse_numeric(item, base)
        return row, diagnostics + parse_diags
    choices, _, _ = extract_choice_sets(item)
    if choices:
        row, parse_diags = parse_multiple_choice(item, base)
        return row, diagnostics + parse_diags
    row, parse_diags = parse_short_or_fitb(item, base)
    return row, diagnostics + parse_diags


def parse_manifest(export_root: Optional[Path]) -> Dict[str, Any]:
    info = {"course_title": "", "quiz_titles_by_file": {}}
    if not export_root or not (export_root / "imsmanifest.xml").exists():
        return info
    manifest = parse_xml(export_root / "imsmanifest.xml")
    title = manifest.findtext(".//{*}lom/{*}general/{*}title/{*}langstring", default="")
    info["course_title"] = title.strip()
    for resource in manifest.findall(".//{*}resource"):
        material_type = attr_local(resource, "material_type", "")
        href = resource.attrib.get("href", "").replace("\\", "/")
        resource_title = resource.attrib.get("title", "").strip()
        if material_type == "d2lquiz" and href:
            info["quiz_titles_by_file"][Path(href).name] = resource_title
    return info


def quiz_to_section_relation_score(quiz_title: str, section_title: str) -> int:
    quiz_tokens = set(normalize_text(quiz_title).split())
    section_tokens = set(normalize_text(section_title).split())
    if not quiz_tokens or not section_tokens:
        return 0
    score = 0
    if "midterm" in quiz_tokens and "midterm" in section_tokens:
        score += 4
    if "final" in quiz_tokens and "final" in section_tokens:
        score += 4
    if "week" in quiz_tokens and "week" in section_tokens:
        score += 2
    score += min(len((quiz_tokens & section_tokens) - {"week", "quiz", "assessment", "exam", "assignment", "timed", "show", "what", "you", "know"}), 4)
    digits = {token for token in quiz_tokens if token.isdigit()}
    if digits and digits & section_tokens:
        score += 3
    return score


class QuestionBank:
    def __init__(self) -> None:
        self.records: List[Dict[str, Any]] = []
        self.records_by_exact_key: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        self.section_items: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        self.sections: Dict[str, Dict[str, str]] = {}

    def add_item(
        self,
        row: Dict[str, Any],
        item_elem: ET.Element,
        section_ident: str,
        section_title: str,
        diagnostics: List[Dict[str, str]],
    ) -> None:
        metadata = qti_metadata(item_elem)
        record = {
            "row": dict(row),
            "question_id": row["question_id"],
            "question_title": row["question_title"],
            "question_type": row["question_type"],
            "section_ident": section_ident,
            "section_title": section_title,
            "stem_norm": normalize_text(row["stem_text"]),
            "title_norm": normalize_text(row["question_title"]),
            "diagnostics": [dict(diag) for diag in diagnostics],
            "metadata": metadata,
        }
        self.records.append(record)
        self.section_items[section_ident].append(record)
        self.sections[section_ident] = {"ident": section_ident, "title": section_title}

        exact_keys = unique_preserve(
            [
                item_elem.attrib.get("label", ""),
                item_elem.attrib.get("ident", ""),
                attr_local(item_elem, "id", ""),
                metadata.get("qmd_displayid", ""),
                metadata.get("qmd_globalid", ""),
            ]
        )
        for key in exact_keys:
            self.records_by_exact_key[key].append(record)

    def match_inline_item(self, item_elem: ET.Element, row: Dict[str, Any], quiz_title: str) -> Dict[str, Any]:
        metadata = qti_metadata(item_elem)
        exact_candidates: List[Dict[str, Any]] = []
        for key in unique_preserve(
            [
                item_elem.attrib.get("label", ""),
                item_elem.attrib.get("ident", ""),
                attr_local(item_elem, "id", ""),
                metadata.get("qmd_displayid", ""),
                metadata.get("qmd_globalid", ""),
            ]
        ):
            exact_candidates.extend(self.records_by_exact_key.get(key, []))

        exact_candidates = list({id(candidate): candidate for candidate in exact_candidates}.values())
        if len(exact_candidates) == 1:
            return {
                "status": "matched",
                "record": exact_candidates[0],
                "reason": "exact stable key match",
            }
        if len(exact_candidates) > 1:
            return {
                "status": "ambiguous",
                "records": exact_candidates,
                "reason": "multiple exact stable key matches",
            }

        title_norm = normalize_text(row["question_title"])
        stem_norm = normalize_text(row["stem_text"])
        question_type = row["question_type"]
        if not title_norm and not stem_norm:
            return {"status": "none"}

        scored = []
        for record in self.records:
            if question_type and record["question_type"] != question_type:
                continue
            score = 0
            reasons = []
            if title_norm and record["title_norm"] and title_norm == record["title_norm"]:
                score += 6
                reasons.append("question title")
            if stem_norm and record["stem_norm"] and stem_norm == record["stem_norm"]:
                score += 5
                reasons.append("stem text")
            if stem_norm and record["stem_norm"] and len(stem_norm) > 20 and stem_norm in record["stem_norm"]:
                score += 2
                reasons.append("stem containment")
            relation_score = quiz_to_section_relation_score(quiz_title, record["section_title"])
            if relation_score:
                score += relation_score
                reasons.append("quiz-to-bank section relation")
            if score >= 8:
                scored.append((score, record, reasons))

        if not scored:
            return {"status": "none"}

        scored.sort(key=lambda item: item[0], reverse=True)
        top_score, top_record, top_reasons = scored[0]
        if len(scored) == 1 or top_score >= scored[1][0] + 2:
            return {
                "status": "matched",
                "record": top_record,
                "reason": ", ".join(top_reasons),
            }
        return {
            "status": "ambiguous",
            "records": [record for score, record, _reasons in scored if score >= top_score - 1],
            "reason": "multiple bank candidates had similar title/stem evidence",
        }

    def resolve_bank_section(self, section_ident: str, section_title: str, quiz_title: str) -> Dict[str, Any]:
        if section_ident in self.section_items:
            return {
                "status": "matched",
                "section_ident": section_ident,
                "section_title": self.sections[section_ident]["title"],
                "records": list(self.section_items[section_ident]),
                "reason": "exact section identifier match",
            }

        title_norm = normalize_text(section_title)
        scored = []
        for candidate_ident, section in self.sections.items():
            score = 0
            reasons = []
            if title_norm and title_norm == normalize_text(section["title"]):
                score += 6
                reasons.append("section title")
            relation_score = quiz_to_section_relation_score(quiz_title, section["title"])
            if relation_score:
                score += relation_score
                reasons.append("quiz-to-bank section relation")
            if score >= 4:
                scored.append((score, candidate_ident, reasons))

        if not scored:
            return {"status": "none"}
        scored.sort(key=lambda item: item[0], reverse=True)
        top_score, top_ident, top_reasons = scored[0]
        if len(scored) == 1 or top_score >= scored[1][0] + 2:
            return {
                "status": "matched",
                "section_ident": top_ident,
                "section_title": self.sections[top_ident]["title"],
                "records": list(self.section_items[top_ident]),
                "reason": ", ".join(top_reasons),
            }
        return {
            "status": "ambiguous",
            "section_idents": [candidate_ident for score, candidate_ident, _ in scored if score >= top_score - 1],
            "reason": "multiple bank sections had similar match scores",
        }


def parse_questiondb(root: Optional[Path], file_index: Optional[Dict[str, Any]] = None) -> Tuple[QuestionBank, List[Dict[str, str]]]:
    bank = QuestionBank()
    source_map: List[Dict[str, str]] = []
    if not root or not (root / "questiondb.xml").exists():
        return bank, source_map

    questiondb = parse_xml(root / "questiondb.xml")
    objectbank = questiondb.find(".//objectbank")
    if objectbank is None:
        return bank, source_map

    for section in objectbank.findall("./section"):
        section_ident = section.attrib.get("ident", "")
        section_title = section.attrib.get("title", "")
        source_map.append(
            make_source_map_row(
                object_type="questiondb_section",
                object_id=section_ident,
                object_title=section_title,
                quiz_id="",
                quiz_title="",
                source_file="questiondb.xml",
                source_hint="objectbank/section",
                resolved_to_sheet="pool_members",
                resolved_to_key=section_ident,
            )
        )
        for item in section.findall("./item"):
            row, diagnostics = parse_item(
                item,
                "questiondb",
                "",
                "questiondb.xml",
                f"objectbank/section[{section_ident}]/item",
            )
            row["source_hint"] = f"objectbank/section[{section_ident}]/item"
            if file_index is not None:
                diagnostics.extend(populate_row_image_fields(row, file_index=file_index, source_file="questiondb.xml"))
            bank.add_item(row, item, section_ident, section_title, diagnostics)
            source_map.append(
                make_source_map_row(
                    object_type="questiondb_item",
                    object_id=row["question_id"],
                    object_title=row["question_title"],
                    quiz_id="",
                    quiz_title="",
                    source_file="questiondb.xml",
                    source_hint=f"objectbank/section[{section_ident}]/item",
                    resolved_to_sheet="questions",
                    resolved_to_key=row["question_id"],
                )
            )
    return bank, source_map


def get_child_sections(container: Optional[ET.Element]) -> List[ET.Element]:
    if container is None:
        return []
    return [child for child in list(container) if local_name(child.tag) == "section"]


def copy_bank_record(record: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    return dict(record["row"]), [dict(diag) for diag in record["diagnostics"]]


def annotate_missing_assets(
    export_root: Path,
    row: Dict[str, Any],
    source_file: str,
) -> List[Dict[str, str]]:
    diagnostics: List[Dict[str, str]] = []
    refs = [ref.strip() for ref in row.get("asset_refs", "").split(";") if ref.strip()]
    missing = []
    for ref in refs:
        cleaned = ref.replace("\\", "/").split("?", 1)[0]
        if not (export_root / cleaned).exists():
            missing.append(ref)
    if missing:
        diagnostics.append(
            diagnostic_seed(
                "missing_assets",
                f"Referenced asset files were not found: {', '.join(missing)}",
                question_id=row["question_id"],
                source_file=source_file,
                source_hint=row.get("source_hint", ""),
                suggested_action="Keep the original Brightspace export bundle with this review, or verify asset paths manually.",
            )
        )
    return diagnostics


def calculate_section_points(
    section_type: str,
    draw_count: str,
    question_rows: List[Dict[str, Any]],
    *,
    quiz_id: str,
    quiz_title: str,
    section_id: str,
    source_file: str,
) -> Tuple[Optional[float], List[Dict[str, str]]]:
    diagnostics: List[Dict[str, str]] = []
    point_values = [parse_number(row.get("points")) for row in question_rows if parse_number(row.get("points")) is not None]
    if not point_values:
        return None, diagnostics

    if section_type != "pool":
        return sum(point_values), diagnostics

    draw_value = parse_number(draw_count)
    distinct_points = sorted(set(point_values))
    if draw_value is not None and len(distinct_points) == 1:
        return distinct_points[0] * draw_value, diagnostics

    diagnostics.append(
        fill_diagnostic_context(
            diagnostic_seed(
                "pool_points_best_effort",
                "Pool points were calculated from candidate questions because draw-count scoring was not uniquely resolvable.",
                source_file=source_file,
                source_hint=f"assessment/section[{section_id}]",
                suggested_action="Verify pool scoring manually if exact attempt totals matter.",
            ),
            quiz_id=quiz_id,
            quiz_title=quiz_title,
            section_id=section_id,
        )
    )
    return sum(point_values), diagnostics


def derive_storage_type(question_rows: List[Dict[str, Any]], diagnostics: List[Dict[str, str]], has_questiondb: bool) -> str:
    locations = {row["source_location"] for row in question_rows if row.get("source_location")}
    if "unresolved" in locations:
        return "unresolved"
    if "questiondb" in locations and "inline" not in locations and "hybrid" not in locations:
        return "banked"
    if "hybrid" in locations or ("questiondb" in locations and ("inline" in locations or has_questiondb)):
        return "hybrid"
    if not question_rows and any(diag["issue_type"].startswith("bank_") for diag in diagnostics):
        return "unresolved"
    return "inline"


def analyze_quiz(
    quiz_path: Path,
    bank: QuestionBank,
    manifest_info: Dict[str, Any],
    file_index: Dict[str, Any],
) -> Dict[str, Any]:
    root = parse_xml(quiz_path)
    assessment = root.find(".//assessment")
    quiz_id = quiz_path.stem.replace("quiz_d2l_", "")
    manifest_title = manifest_info["quiz_titles_by_file"].get(quiz_path.name, "")
    quiz_title = ""
    if assessment is not None:
        quiz_title = assessment.attrib.get("title", "").strip()
    quiz_title = quiz_title or manifest_title or quiz_path.name
    instructions_html = assessment.findtext("./rubric/flow_mat/material/mattext", default="") if assessment is not None else ""

    questions_rows: List[Dict[str, Any]] = []
    section_rows: List[Dict[str, Any]] = []
    pool_rows: List[Dict[str, Any]] = []
    diagnostics: List[Dict[str, str]] = []
    source_map: List[Dict[str, str]] = []

    declared_total_points = 0.0
    resolved_total_points = 0.0
    section_order = 0
    pool_count = 0

    container = assessment.find("./section") if assessment is not None else None
    sections = get_child_sections(container)
    if not sections and container is not None:
        sections = [container]

    for section in sections:
        section_order += 1
        section_ident = section.attrib.get("ident", f"SECT_{section_order}")
        section_title = section.attrib.get("title", "") or ("Quiz Section" if section_ident.startswith("RAND_") else f"Section {section_order}")
        metadata = qti_metadata(section)
        section_type = "pool" if section_ident.startswith("RAND_") or metadata.get("qmd_numberofitems") else "section"
        draw_count = metadata.get("qmd_numberofitems", "") if section_type == "pool" else ""
        items = section.findall("./item")

        section_row = {
            "quiz_id": quiz_id,
            "quiz_title": quiz_title,
            "section_id": section_ident,
            "section_title": section_title,
            "section_order": section_order,
            "section_type": section_type,
            "draw_count": draw_count,
            "pool_size": "",
            "section_points_total": "",
            "question_count": 0,
            "source_quiz_file": quiz_path.name,
            "source_bank_file": "questiondb.xml" if bank.records else "",
            "notes": "",
        }
        section_rows.append(section_row)
        source_map.append(
            make_source_map_row(
                object_type="quiz_section",
                object_id=section_ident,
                object_title=section_title,
                quiz_id=quiz_id,
                quiz_title=quiz_title,
                source_file=quiz_path.name,
                source_hint=f"assessment/section[{section_ident}]",
                resolved_to_sheet="sections_pools",
                resolved_to_key=section_ident,
            )
        )

        resolved_section_questions: List[Dict[str, Any]] = []
        resolved_section_diags: List[Dict[str, str]] = []

        if items:
            for question_order, item in enumerate(items, start=1):
                row, item_diags = parse_item(
                    item,
                    "inline",
                    quiz_path.name,
                    "",
                    f"assessment/section[{section_ident}]/item",
                )
                row.update(
                    {
                        "quiz_id": quiz_id,
                        "quiz_title": quiz_title,
                        "section_id": section_ident,
                        "section_title": section_title,
                        "question_order": question_order,
                    }
                )
                item_diags.extend(populate_row_image_fields(row, file_index=file_index, source_file=quiz_path.name))

                match = bank.match_inline_item(item, row, quiz_title) if bank.records else {"status": "none"}
                if match["status"] == "matched":
                    record = match["record"]
                    row["source_location"] = "hybrid"
                    row["source_bank_file"] = "questiondb.xml"
                    row["source_hint"] = (
                        f"assessment/section[{section_ident}]/item; "
                        f"matched questiondb section {record['section_ident']} via {match['reason']}"
                    )
                    source_map.append(
                        make_source_map_row(
                            object_type="hybrid_bank_match",
                            object_id=row["question_id"],
                            object_title=row["question_title"],
                            quiz_id=quiz_id,
                            quiz_title=quiz_title,
                            source_file="questiondb.xml",
                            source_hint=f"matched questiondb section {record['section_ident']} via {match['reason']}",
                            resolved_to_sheet="questions",
                            resolved_to_key=row["question_id"],
                        )
                    )
                elif match["status"] == "ambiguous":
                    item_diags.append(
                        diagnostic_seed(
                            "ambiguous_bank_match",
                            "Inline question had multiple plausible questiondb matches; source remains inline.",
                            question_id=row["question_id"],
                            source_file=quiz_path.name,
                            source_hint=row["source_hint"],
                            suggested_action="Review the question title/stem and question bank entries manually.",
                        )
                    )
                row_diags = [
                    fill_diagnostic_context(
                        diag,
                        quiz_id=quiz_id,
                        quiz_title=quiz_title,
                        section_id=section_ident,
                        question_id=row["question_id"],
                    )
                    for diag in item_diags
                ]
                row_diags.extend(
                    fill_diagnostic_context(
                        diag,
                        quiz_id=quiz_id,
                        quiz_title=quiz_title,
                        section_id=section_ident,
                        question_id=row["question_id"],
                    )
                    for diag in annotate_missing_assets(quiz_path.parent, row, quiz_path.name)
                )
                resolved_section_diags.extend(row_diags)
                resolved_section_questions.append(row)
                source_map.append(
                    make_source_map_row(
                        object_type="quiz_item",
                        object_id=row["question_id"],
                        object_title=row["question_title"],
                        quiz_id=quiz_id,
                        quiz_title=quiz_title,
                        source_file=quiz_path.name,
                        source_hint=row["source_hint"],
                        resolved_to_sheet="questions",
                        resolved_to_key=row["question_id"],
                    )
                )
        elif bank.records and section_type == "pool":
            section_match = bank.resolve_bank_section(section_ident, section_title, quiz_title)
            if section_match["status"] == "matched":
                section_row["notes"] = f"Bank-backed pool resolved via {section_match['reason']}"
                for question_order, record in enumerate(section_match["records"], start=1):
                    row, item_diags = copy_bank_record(record)
                    row.update(
                        {
                            "quiz_id": quiz_id,
                            "quiz_title": quiz_title,
                            "section_id": section_ident,
                            "section_title": section_title,
                            "question_order": question_order,
                            "source_location": "questiondb",
                            "source_quiz_file": quiz_path.name,
                            "source_bank_file": "questiondb.xml",
                            "source_hint": (
                                f"assessment/section[{section_ident}] resolved to questiondb section "
                                f"{section_match['section_ident']} via {section_match['reason']}"
                            ),
                        }
                    )
                    row_diags = [
                        fill_diagnostic_context(
                            diag,
                            quiz_id=quiz_id,
                            quiz_title=quiz_title,
                            section_id=section_ident,
                            question_id=row["question_id"],
                        )
                        for diag in item_diags
                    ]
                    resolved_section_diags.extend(row_diags)
                    resolved_section_questions.append(row)
                    source_map.append(
                        make_source_map_row(
                            object_type="bank_only_pool_item",
                            object_id=row["question_id"],
                            object_title=row["question_title"],
                            quiz_id=quiz_id,
                            quiz_title=quiz_title,
                            source_file="questiondb.xml",
                            source_hint=row["source_hint"],
                            resolved_to_sheet="questions",
                            resolved_to_key=row["question_id"],
                        )
                    )
            elif section_match["status"] == "ambiguous":
                diagnostics.append(
                    fill_diagnostic_context(
                        diagnostic_seed(
                            "bank_only_ambiguous",
                            "Pool/section appears bank-backed but multiple questiondb sections matched with similar confidence.",
                            source_file=quiz_path.name,
                            source_hint=f"assessment/section[{section_ident}]",
                            suggested_action="Review pool-to-bank mapping manually.",
                        ),
                        quiz_id=quiz_id,
                        quiz_title=quiz_title,
                        section_id=section_ident,
                    )
                )
                pool_rows.append(
                    {
                        "quiz_id": quiz_id,
                        "quiz_title": quiz_title,
                        "section_id": section_ident,
                        "pool_title": section_title,
                        "pool_draw_count": draw_count,
                        "pool_size": "",
                        "question_id": "",
                        "question_title": "(unresolved bank-backed pool)",
                        "question_type": "",
                        "points": "",
                        "source_location": "unresolved",
                        "source_hint": "Multiple questiondb sections matched this pool.",
                        "source_quiz_file": quiz_path.name,
                        "source_bank_file": "questiondb.xml",
                        "included_in_resolved_review": "no",
                        "reviewer_notes": "",
                    }
                )
            else:
                diagnostics.append(
                    fill_diagnostic_context(
                        diagnostic_seed(
                            "bank_only_unresolved",
                            "Pool/section appears bank-backed but no confident questiondb section mapping was found.",
                            source_file=quiz_path.name,
                            source_hint=f"assessment/section[{section_ident}]",
                            suggested_action="Review section-to-bank mapping manually.",
                        ),
                        quiz_id=quiz_id,
                        quiz_title=quiz_title,
                        section_id=section_ident,
                    )
                )
                pool_rows.append(
                    {
                        "quiz_id": quiz_id,
                        "quiz_title": quiz_title,
                        "section_id": section_ident,
                        "pool_title": section_title,
                        "pool_draw_count": draw_count,
                        "pool_size": "",
                        "question_id": "",
                        "question_title": "(unresolved bank-backed pool)",
                        "question_type": "",
                        "points": "",
                        "source_location": "unresolved",
                        "source_hint": "No confident questiondb section mapping was found.",
                        "source_quiz_file": quiz_path.name,
                        "source_bank_file": "questiondb.xml",
                        "included_in_resolved_review": "no",
                        "reviewer_notes": "",
                    }
                )

        diagnostics.extend(resolved_section_diags)
        for row in resolved_section_questions:
            questions_rows.append(row)
            section_row["question_count"] += 1
            if section_type == "pool":
                pool_rows.append(
                    {
                        "quiz_id": quiz_id,
                        "quiz_title": quiz_title,
                        "section_id": section_ident,
                        "pool_title": section_title,
                        "pool_draw_count": draw_count,
                        "pool_size": "",
                        "question_id": row["question_id"],
                        "question_title": row["question_title"],
                        "question_type": row["question_type"],
                        "points": row["points"],
                        "source_location": row["source_location"],
                        "source_hint": row["source_hint"],
                        "source_quiz_file": quiz_path.name,
                        "source_bank_file": row["source_bank_file"],
                        "included_in_resolved_review": "yes",
                        "reviewer_notes": "",
                    }
                )

        if section_type == "pool":
            pool_count += 1
            section_row["pool_size"] = section_row["question_count"] or ""
            for pool_row in pool_rows:
                if pool_row["quiz_id"] == quiz_id and pool_row["section_id"] == section_ident:
                    pool_row["pool_size"] = section_row["pool_size"]

        section_points_total, point_diags = calculate_section_points(
            section_type,
            draw_count,
            resolved_section_questions,
            quiz_id=quiz_id,
            quiz_title=quiz_title,
            section_id=section_ident,
            source_file=quiz_path.name,
        )
        diagnostics.extend(point_diags)
        section_row["section_points_total"] = format_number(section_points_total)
        if section_points_total is not None:
            declared_total_points += section_points_total
            resolved_total_points += section_points_total

    question_id_counts = Counter(row["question_id"] for row in questions_rows if row.get("question_id"))
    for question_id, count in sorted(question_id_counts.items()):
        if count > 1:
            diagnostics.append(
                fill_diagnostic_context(
                    diagnostic_seed(
                        "duplicate_question_id",
                        f"Question ID '{question_id}' appeared {count} times in this quiz review output.",
                        source_file=quiz_path.name,
                        source_hint="questions",
                        suggested_action="Use quiz title, section, and source hint when reviewing duplicates.",
                    ),
                    quiz_id=quiz_id,
                    quiz_title=quiz_title,
                    question_id=question_id,
                )
            )

    if bank.records and pool_count:
        diagnostics.append(
            fill_diagnostic_context(
                diagnostic_seed(
                    "pool_dependency",
                    "Question bank content is relevant to at least one pooled quiz section; keep questiondb.xml with this review bundle.",
                    source_file=quiz_path.name,
                    source_hint="questiondb.xml",
                    suggested_action="Retain questiondb.xml alongside the quiz review outputs.",
                ),
                quiz_id=quiz_id,
                quiz_title=quiz_title,
            )
        )

    storage_type = derive_storage_type(questions_rows, diagnostics, bool(bank.records))
    delivery_type = "fixed"
    if pool_count and len(section_rows) == pool_count:
        delivery_type = "pooled"
    elif pool_count:
        delivery_type = "mixed"

    attempts_allowed = assessment.findtext(f".//{{{D2L_NS}}}attempts_allowed", default="") if assessment is not None else ""
    time_limit_minutes = assessment.findtext(f".//{{{D2L_NS}}}time_limit", default="") if assessment is not None else ""
    shuffle_questions = assessment.findtext(f".//{{{D2L_NS}}}shuffle_questions", default="") if assessment is not None else ""
    shuffle_answers = assessment.findtext(f".//{{{D2L_NS}}}shuffle_answers", default="") if assessment is not None else ""

    overview = {
        "quiz_id": quiz_id,
        "quiz_title": quiz_title,
        "source_quiz_file": quiz_path.name,
        "storage_type": storage_type,
        "delivery_type": delivery_type,
        "has_questiondb": bool(bank.records),
        "time_limit_minutes": time_limit_minutes,
        "attempts_allowed": attempts_allowed,
        "shuffle_questions": shuffle_questions or "unknown",
        "shuffle_answers": shuffle_answers or "unknown",
        "declared_total_points": format_number(declared_total_points),
        "resolved_total_points": format_number(resolved_total_points),
        "question_count_resolved": len(questions_rows),
        "section_count": len(section_rows),
        "pool_count": pool_count,
        "instructions_text": html_to_text(instructions_html),
        "review_notes": "",
    }

    return {
        "overview": overview,
        "sections": section_rows,
        "questions": questions_rows,
        "pool_members": pool_rows,
        "diagnostics": diagnostics,
        "source_map": source_map,
    }


def autosize(ws, sheet_name: str) -> None:
    headers = [cell.value or "" for cell in ws[1]]
    overrides = TEXT_WIDTH_OVERRIDES.get(sheet_name, {})
    for column_cells in ws.columns:
        column_index = column_cells[0].column
        header = headers[column_index - 1] if column_index - 1 < len(headers) else ""
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        width = min(max(max_length + 2, 12), 55)
        if header in overrides:
            width = overrides[header]
        ws.column_dimensions[get_column_letter(column_index)].width = width


def style_header(ws, sheet_name: str, columns: List[str]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    secondary_fill = PatternFill("solid", fgColor="7D8FA3")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    raw_divider = Side(style="medium", color="7F7F7F")
    raw_start_index = columns.index("question_title") + 1 if sheet_name == "questions" and "question_title" in columns else None
    for cell in ws[1]:
        use_secondary = raw_start_index is not None and cell.column >= raw_start_index
        cell.fill = secondary_fill if use_secondary else fill
        cell.font = font
        cell.border = Border(
            left=raw_divider if raw_start_index is not None and cell.column == raw_start_index else Side(style=None),
            bottom=thin,
        )
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def style_diagnostics_rows(ws, columns: List[str]) -> None:
    if "severity" not in columns:
        return
    severity_column = columns.index("severity") + 1
    fills = {
        "info": PatternFill("solid", fgColor="E2F0D9"),
        "warning": PatternFill("solid", fgColor="FFF2CC"),
        "error": PatternFill("solid", fgColor="FCE4D6"),
    }
    for row in ws.iter_rows(min_row=2):
        severity = (row[severity_column - 1].value or "").lower()
        fill = fills.get(severity)
        if fill is None:
            continue
        row[severity_column - 1].fill = fill


def apply_image_hyperlinks(ws, columns: List[str]) -> None:
    if "image_link_primary" not in columns:
        return
    column_index = columns.index("image_link_primary") + 1
    for row in ws.iter_rows(min_row=2):
        cell = row[column_index - 1]
        if not cell.value:
            continue
        cell.hyperlink = str(cell.value)
        cell.style = "Hyperlink"


def apply_question_group_divider(ws, columns: List[str]) -> None:
    if "question_title" not in columns:
        return
    column_index = columns.index("question_title") + 1
    raw_divider = Side(style="medium", color="7F7F7F")
    for row in ws.iter_rows(min_row=2):
        cell = row[column_index - 1]
        cell.border = Border(
            left=raw_divider,
            right=cell.border.right,
            top=cell.border.top,
            bottom=cell.border.bottom,
        )


def estimate_cell_lines(value: Any, width: float) -> int:
    text = str(value or "")
    if not text:
        return 1
    width_chars = max(int(width * 1.1), 1)
    lines = 0
    for line in text.splitlines() or [""]:
        lines += max(1, math.ceil(len(line) / width_chars))
    return max(lines, 1)


def apply_row_heights(ws, rows: List[Dict[str, Any]], columns: List[str], sheet_name: str) -> None:
    if sheet_name == "questions":
        measure_columns = [
            "question_title_review",
            "stem_text",
            "matching_review_display",
            "ordering_review_display",
            "correct_answer_text",
            "general_feedback",
            "correct_feedback",
            "incorrect_feedback",
            "answer_specific_feedback",
            "grading_notes",
            "reviewer_notes",
        ]
        for row_index, row in enumerate(rows, start=2):
            minimum_height = 15
            question_type = str(row.get("question_type", ""))
            if question_type == "Matching":
                minimum_height = max(minimum_height, 72)
            elif question_type == "Ordering":
                minimum_height = max(minimum_height, 60)
            elif question_type == "Long Answer":
                minimum_height = max(minimum_height, 54)
            if question_type not in {"Matching", "Ordering"} and str(row.get("has_image", "no")).lower() == "yes":
                minimum_height = max(minimum_height, 45)

            max_lines = 1
            for column in measure_columns:
                if column not in columns:
                    continue
                column_index = columns.index(column) + 1
                width = ws.column_dimensions[get_column_letter(column_index)].width or 12
                max_lines = max(max_lines, estimate_cell_lines(row.get(column, ""), width))
            ws.row_dimensions[row_index].height = max(minimum_height, max_lines * 15)
    elif sheet_name == "matching_pairs_expanded":
        measure_columns = ["prompt", "correct_match", "question_title_review", "source_hint"]
        for row_index, row in enumerate(rows, start=2):
            max_lines = 1
            for column in measure_columns:
                if column not in columns:
                    continue
                column_index = columns.index(column) + 1
                width = ws.column_dimensions[get_column_letter(column_index)].width or 12
                max_lines = max(max_lines, estimate_cell_lines(row.get(column, ""), width))
            ws.row_dimensions[row_index].height = max(42, max_lines * 15)


def write_sheet(ws, rows: List[Dict[str, Any]], columns: List[str], sheet_name: str) -> None:
    ws.append(columns)
    style_header(ws, sheet_name, columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    autosize(ws, sheet_name)
    apply_row_heights(ws, rows, columns, sheet_name)
    if sheet_name == "diagnostics":
        style_diagnostics_rows(ws, columns)
    apply_image_hyperlinks(ws, columns)
    if sheet_name == "questions":
        apply_question_group_divider(ws, columns)


def build_workbook() -> Workbook:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for sheet_name in SHEET_ORDER:
        workbook.create_sheet(sheet_name)
    return workbook


SHEET_SPECS = {
    "quiz_overview": [
        "quiz_id",
        "quiz_title",
        "source_quiz_file",
        "storage_type",
        "delivery_type",
        "has_questiondb",
        "time_limit_minutes",
        "attempts_allowed",
        "shuffle_questions",
        "shuffle_answers",
        "declared_total_points",
        "resolved_total_points",
        "question_count_resolved",
        "section_count",
        "pool_count",
        "instructions_text",
        "review_notes",
    ],
    "sections_pools": [
        "quiz_id",
        "quiz_title",
        "section_id",
        "section_title",
        "section_order",
        "section_type",
        "draw_count",
        "pool_size",
        "section_points_total",
        "question_count",
        "source_quiz_file",
        "source_bank_file",
        "notes",
    ],
    "questions": [
        "quiz_id",
        "quiz_title",
        "section_id",
        "section_title",
        "question_order",
        "question_id",
        "question_title_review",
        "question_type",
        "points",
        "has_image",
        "image_count",
        "image_link_primary",
        "stem_text",
        "matching_review_display",
        "ordering_review_display",
        "choice_a",
        "choice_b",
        "choice_c",
        "choice_d",
        "choice_e",
        "choice_f",
        "correct_answer_key",
        "correct_answer_text",
        "all_correct_keys",
        "accepted_answers",
        "matching_pairs",
        "ordering_sequence",
        "numeric_answer",
        "numeric_tolerance",
        "response_schema",
        "general_feedback",
        "correct_feedback",
        "incorrect_feedback",
        "answer_specific_feedback",
        "grading_notes",
        "review_status",
        "reviewer_notes",
        "question_title",
        "source_location",
        "source_hint",
        "source_quiz_file",
        "source_bank_file",
        "response_schema",
        "metadata_difficulty",
        "metadata_mandatory",
        "all_correct_keys",
        "accepted_answers",
        "matching_pairs",
        "ordering_sequence",
        "numeric_answer",
        "numeric_tolerance",
        "asset_refs",
        "image_refs",
        "image_paths_resolved",
        "question_payload_json",
    ],
    "matching_pairs_expanded": [
        "quiz_id",
        "quiz_title",
        "section_id",
        "section_title",
        "question_order",
        "question_id",
        "question_title_review",
        "question_title",
        "prompt_order",
        "match_order",
        "prompt",
        "correct_match",
        "has_image",
        "image_count",
        "image_refs",
        "image_paths_resolved",
        "image_link_primary",
        "source_location",
        "source_quiz_file",
        "source_bank_file",
        "source_hint",
    ],
    "pool_members": [
        "quiz_id",
        "quiz_title",
        "section_id",
        "pool_title",
        "pool_draw_count",
        "pool_size",
        "question_id",
        "question_title",
        "question_type",
        "points",
        "source_location",
        "source_hint",
        "source_quiz_file",
        "source_bank_file",
        "included_in_resolved_review",
        "reviewer_notes",
    ],
    "diagnostics": [
        "severity",
        "quiz_id",
        "quiz_title",
        "section_id",
        "question_id",
        "issue_type",
        "message",
        "source_file",
        "source_hint",
        "suggested_action",
    ],
    "source_map": [
        "object_type",
        "object_id",
        "object_title",
        "quiz_id",
        "quiz_title",
        "source_file",
        "source_hint",
        "resolved_to_sheet",
        "resolved_to_key",
    ],
}


def add_template_examples(workbook: Workbook) -> None:
    examples = {
        "quiz_overview": [
            {
                "quiz_id": "5204",
                "quiz_title": "Final Exam",
                "source_quiz_file": "quiz_d2l_5204.xml",
                "storage_type": "hybrid",
                "delivery_type": "mixed",
                "has_questiondb": True,
                "time_limit_minutes": 120,
                "attempts_allowed": 1,
                "shuffle_questions": "TRUE",
                "shuffle_answers": "FALSE",
                "declared_total_points": 100,
                "resolved_total_points": 100,
                "question_count_resolved": 50,
                "section_count": 6,
                "pool_count": 2,
                "instructions_text": "Timed exam. Example row.",
                "review_notes": "EXAMPLE",
            }
        ],
        "sections_pools": [
            {
                "quiz_id": "5204",
                "quiz_title": "Final Exam",
                "section_id": "POOL_02",
                "section_title": "Random Diagnostic Cases",
                "section_order": 4,
                "section_type": "pool",
                "draw_count": 5,
                "pool_size": 20,
                "section_points_total": 25,
                "question_count": 20,
                "source_quiz_file": "quiz_d2l_5204.xml",
                "source_bank_file": "questiondb.xml",
                "notes": "EXAMPLE",
            }
        ],
        "questions": [
            {
                "quiz_id": "5204",
                "quiz_title": "Final Exam",
                "section_id": "SECT_01",
                "section_title": "Cardiology",
                "question_order": 1,
                "question_id": "QUES_0001",
                "question_title": "Question 1",
                "question_title_review": "Question 1",
                "question_type": "Multiple Choice",
                "points": 2,
                "has_image": "yes",
                "image_count": 1,
                "image_link_primary": "assets/csfiles/home_dir/example_heart.png",
                "stem_text": "Which vessel returns oxygenated blood from the lungs to the heart?",
                "matching_review_display": "",
                "ordering_review_display": "",
                "choice_a": "Aorta",
                "choice_b": "Pulmonary vein",
                "choice_c": "Vena cava",
                "choice_d": "Pulmonary artery",
                "correct_answer_key": "B",
                "correct_answer_text": "Pulmonary vein",
                "general_feedback": "Review the circulatory loop if needed.",
                "correct_feedback": "",
                "incorrect_feedback": "",
                "answer_specific_feedback": "",
                "grading_notes": "",
                "review_status": "pending",
                "reviewer_notes": "EXAMPLE",
                "source_location": "hybrid",
                "source_hint": "assessment/section[SECT_01]/item; matched questiondb section SECT_48305 via question title, quiz-to-bank section relation",
                "source_quiz_file": "quiz_d2l_5204.xml",
                "source_bank_file": "questiondb.xml",
                "response_schema": "single_select",
                "metadata_difficulty": "2",
                "metadata_mandatory": "yes",
                "all_correct_keys": "B",
                "accepted_answers": "",
                "matching_pairs": "",
                "ordering_sequence": "",
                "numeric_answer": "",
                "numeric_tolerance": "",
                "asset_refs": "",
                "image_refs": "csfiles/home_dir/example_heart.png",
                "image_paths_resolved": "csfiles/home_dir/example_heart.png",
                "question_payload_json": "",
            }
        ],
        "matching_pairs_expanded": [
            {
                "quiz_id": "5204",
                "quiz_title": "Final Exam",
                "section_id": "SECT_03",
                "section_title": "Labeling",
                "question_order": 4,
                "question_id": "QUES_0042",
                "question_title_review": "Match the anatomy labels to the correct structure.",
                "question_title": "QUES_0042",
                "prompt_order": 1,
                "match_order": 1,
                "prompt": "A",
                "correct_match": "Left ventricle",
                "has_image": "yes",
                "image_count": 1,
                "image_refs": "csfiles/home_dir/example_heart.png",
                "image_paths_resolved": "csfiles/home_dir/example_heart.png",
                "image_link_primary": "assets/csfiles/home_dir/example_heart.png",
                "source_location": "inline",
                "source_quiz_file": "quiz_d2l_5204.xml",
                "source_bank_file": "",
                "source_hint": "assessment/section[SECT_03]/item",
            }
        ],
        "pool_members": [
            {
                "quiz_id": "5204",
                "quiz_title": "Final Exam",
                "section_id": "POOL_02",
                "pool_title": "Random Diagnostic Cases",
                "pool_draw_count": 5,
                "pool_size": 20,
                "question_id": "QUES_1044",
                "question_title": "Case 7",
                "question_type": "Multiple Choice",
                "points": 5,
                "source_location": "questiondb",
                "source_hint": "assessment/section[POOL_02] resolved to questiondb section SECT_48305 via quiz-to-bank section relation",
                "source_quiz_file": "quiz_d2l_5204.xml",
                "source_bank_file": "questiondb.xml",
                "included_in_resolved_review": "yes",
                "reviewer_notes": "EXAMPLE",
            }
        ],
        "diagnostics": [
            {
                "severity": "warning",
                "quiz_id": "5204",
                "quiz_title": "Final Exam",
                "section_id": "POOL_02",
                "question_id": "",
                "issue_type": "pool_dependency",
                "message": "Question bank content is relevant to at least one pooled quiz section; keep questiondb.xml with this review bundle.",
                "source_file": "quiz_d2l_5204.xml",
                "source_hint": "questiondb.xml",
                "suggested_action": "Retain questiondb.xml alongside the quiz review outputs.",
            }
        ],
        "source_map": [
            {
                "object_type": "quiz_item",
                "object_id": "QUES_0001",
                "object_title": "Question 1",
                "quiz_id": "5204",
                "quiz_title": "Final Exam",
                "source_file": "quiz_d2l_5204.xml",
                "source_hint": "assessment/section[SECT_01]/item",
                "resolved_to_sheet": "questions",
                "resolved_to_key": "QUES_0001",
            }
        ],
    }

    for sheet_name, columns in SHEET_SPECS.items():
        write_sheet(workbook[sheet_name], examples.get(sheet_name, []), columns, sheet_name)


def create_review_outputs(export_root: Path, out_dir: Path, copy_images_to_assets: bool = False) -> Path:
    manifest_info = parse_manifest(export_root)
    file_index = build_export_file_index(export_root)
    bank, questiondb_source_map = parse_questiondb(export_root, file_index)
    quiz_results = []
    sections = []
    questions = []
    pools = []
    diagnostics = []
    source_map = list(questiondb_source_map)

    for quiz_path in sorted(export_root.glob("quiz_d2l_*.xml")):
        result = analyze_quiz(quiz_path, bank, manifest_info, file_index)
        quiz_results.append(result["overview"])
        sections.extend(result["sections"])
        questions.extend(result["questions"])
        pools.extend(result["pool_members"])
        diagnostics.extend(result["diagnostics"])
        source_map.extend(result["source_map"])

    enrich_question_rows_for_review(questions)
    diagnostics.extend(
        finalize_question_image_links(
            questions,
            export_root=export_root,
            out_dir=out_dir,
            copy_images_to_assets=copy_images_to_assets,
        )
    )
    matching_pairs_expanded = build_matching_pairs_expanded_rows(questions)

    workbook = build_workbook()
    row_map = {
        "quiz_overview": quiz_results,
        "sections_pools": sections,
        "questions": questions,
        "matching_pairs_expanded": matching_pairs_expanded,
        "pool_members": pools,
        "diagnostics": diagnostics,
        "source_map": source_map,
    }
    for sheet_name, columns in SHEET_SPECS.items():
        write_sheet(workbook[sheet_name], row_map[sheet_name], columns, sheet_name)

    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / "quiz_review.xlsx"
    workbook.save(xlsx_path)

    data = {
        "quiz_overview": quiz_results,
        "sections_pools": sections,
        "questions": questions,
        "matching_pairs_expanded": matching_pairs_expanded,
        "pool_members": pools,
        "diagnostics": diagnostics,
        "source_map": source_map,
    }
    (out_dir / "quiz_review.json").write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

    diagnostics_by_quiz: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for diag in diagnostics:
        diagnostics_by_quiz[diag["quiz_id"]].append(diag)
    questions_by_quiz: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in questions:
        questions_by_quiz[row["quiz_id"]].append(row)

    lines = ["# Brightspace Quiz Review Summary", ""]
    for quiz in quiz_results:
        quiz_diags = diagnostics_by_quiz.get(quiz["quiz_id"], [])
        quiz_questions = questions_by_quiz.get(quiz["quiz_id"], [])
        issue_counts = Counter(diag["issue_type"] for diag in quiz_diags)
        top_issues = ", ".join(f"{issue_type} ({count})" for issue_type, count in issue_counts.most_common(3))
        lines.extend(
            [
                f"## {quiz['quiz_title']}",
                f"- Quiz ID: {quiz['quiz_id']}",
                f"- Source file: {quiz['source_quiz_file']}",
                f"- Storage type: {quiz['storage_type']}",
                f"- Delivery type: {quiz['delivery_type']}",
                f"- Question rows: {quiz['question_count_resolved']}",
                f"- Sections: {quiz['section_count']}",
                f"- Pools: {quiz['pool_count']}",
                f"- Declared / resolved points: {quiz['declared_total_points']} / {quiz['resolved_total_points']}",
                f"- Diagnostics: {len(quiz_diags)}",
            ]
        )
        image_questions = sum(1 for row in quiz_questions if split_semicolon_values(row.get("image_refs", "")))
        resolved_images = sum(int(row.get("image_count") or 0) for row in quiz_questions)
        image_diags = sum(1 for diag in quiz_diags if "image" in diag["issue_type"])
        if image_questions or image_diags:
            lines.append(
                f"- Images: {image_questions} question(s), {resolved_images} resolved image(s), {image_diags} image diagnostic(s)"
            )
        if top_issues:
            lines.append(f"- Major issues: {top_issues}")
        lines.append("")
    (out_dir / "quiz_review_summary.md").write_text("\n".join(lines), encoding="utf-8")
    return xlsx_path


def resolve_input(input_path: Path) -> Tuple[Path, Optional[tempfile.TemporaryDirectory]]:
    if input_path.is_dir():
        return input_path, None
    if zipfile.is_zipfile(input_path):
        temp_dir = tempfile.TemporaryDirectory(prefix="quiz_review_")
        with zipfile.ZipFile(input_path) as archive:
            archive.extractall(temp_dir.name)
        root = Path(temp_dir.name)
        manifest = next(root.rglob("imsmanifest.xml"), None)
        if manifest is None:
            raise SystemExit("Could not find imsmanifest.xml in input zip")
        return manifest.parent, temp_dir
    raise SystemExit("Input must be a Brightspace export zip or unpacked export folder")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build a Brightspace quiz-review workbook, JSON, and Markdown summary from an export."
    )
    parser.add_argument("input", nargs="?", help="Brightspace export ZIP or unpacked folder")
    parser.add_argument("--out", default="quiz_review_out", help="Output directory")
    parser.add_argument("--template-only", action="store_true", help="Create only the workbook template")
    parser.add_argument(
        "--copy-images-to-assets",
        action="store_true",
        help="Copy resolved question images into an assets/ folder beside the workbook and link to the copied files.",
    )
    args = parser.parse_args()

    out_dir = Path(args.out).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.template_only:
        workbook = build_workbook()
        add_template_examples(workbook)
        output_path = out_dir / "quiz_review_template.xlsx"
        workbook.save(output_path)
        print(output_path)
        return

    if not args.input:
        raise SystemExit("Input ZIP/folder required unless using --template-only")

    export_root, temp_dir = resolve_input(Path(args.input).expanduser().resolve())
    try:
        xlsx_path = create_review_outputs(export_root, out_dir, copy_images_to_assets=args.copy_images_to_assets)
        print(xlsx_path)
    finally:
        if temp_dir is not None:
            temp_dir.cleanup()


if __name__ == "__main__":
    main()
