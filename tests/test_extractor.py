from __future__ import annotations

import importlib.util
import json
from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
import pytest

REPO_ROOT = Path(__file__).resolve().parents[1]
MODULE_PATH = REPO_ROOT / "brightspace_quiz_review_extractor_v2.py"
SPEC = importlib.util.spec_from_file_location("brightspace_quiz_review_extractor_v2", MODULE_PATH)
extractor = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
SPEC.loader.exec_module(extractor)


def item_from_xml(xml_text: str) -> ET.Element:
    return ET.fromstring(extractor.sanitize_xml(xml_text))


def build_questiondb_xml(section_blocks: str) -> str:
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<questestinterop>
  <objectbank ident="BANK_1" title="Question Library">
    {section_blocks}
  </objectbank>
</questestinterop>
"""


def write_questiondb(tmp_path: Path, section_blocks: str) -> Path:
    questiondb = tmp_path / "questiondb.xml"
    questiondb.write_text(build_questiondb_xml(section_blocks), encoding="utf-8")
    return questiondb


def build_quiz_xml(item_block: str, *, quiz_title: str = "Image Quiz", quiz_ident: str = "res_quiz_100") -> str:
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<questestinterop xmlns:d2l_2p0="http://desire2learn.com/xsd/d2lcp_v2p0">
  <assessment ident="{quiz_ident}" title="{quiz_title}">
    <section ident="SECT_1" title="Section 1">
      {item_block}
    </section>
  </assessment>
</questestinterop>
"""


def build_manifest_xml(*, quiz_file: str = "quiz_d2l_100.xml", quiz_title: str = "Image Quiz") -> str:
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<manifest xmlns="http://www.imsglobal.org/xsd/imscp_v1p1p2" xmlns:d2l_2p0="http://desire2learn.com/xsd/d2lcp_v2p0">
  <metadata>
    <lom xmlns="http://ltsc.ieee.org/xsd/LOM">
      <general>
        <title><langstring>Sample Course</langstring></title>
      </general>
    </lom>
  </metadata>
  <resources>
    <resource identifier="RES_1" href="{quiz_file}" title="{quiz_title}" d2l_2p0:material_type="d2lquiz" />
  </resources>
</manifest>
"""


def create_minimal_export(tmp_path: Path, item_block: str, *, quiz_file: str = "quiz_d2l_100.xml") -> Path:
    export_root = tmp_path / "export"
    export_root.mkdir()
    (export_root / "imsmanifest.xml").write_text(build_manifest_xml(quiz_file=quiz_file), encoding="utf-8")
    (export_root / quiz_file).write_text(build_quiz_xml(item_block), encoding="utf-8")
    return export_root


def test_parse_matching_question():
    item = item_from_xml(
        """
<item label="MATCH_1" title="Match Terms">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Matching</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_weighting</fieldlabel><fieldentry>2.000000000</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation>
    <flow>
      <material><mattext texttype="text/html">Match the pairs.</mattext></material>
      <response_grp respident="RESP_1" rcardinality="Single">
        <material><mattext texttype="text/html">First prompt</mattext></material>
        <render_choice shuffle="yes">
          <flow_label class="Block">
            <response_label ident="OPT_A"><flow_mat><material><mattext texttype="text/html">Alpha</mattext></material></flow_mat></response_label>
            <response_label ident="OPT_B"><flow_mat><material><mattext texttype="text/html">Beta</mattext></material></flow_mat></response_label>
          </flow_label>
        </render_choice>
      </response_grp>
      <response_grp respident="RESP_2" rcardinality="Single">
        <material><mattext texttype="text/html">Second prompt</mattext></material>
        <render_choice shuffle="yes">
          <flow_label class="Block">
            <response_label ident="OPT_A"><flow_mat><material><mattext texttype="text/html">Alpha</mattext></material></flow_mat></response_label>
            <response_label ident="OPT_B"><flow_mat><material><mattext texttype="text/html">Beta</mattext></material></flow_mat></response_label>
          </flow_label>
        </render_choice>
      </response_grp>
    </flow>
  </presentation>
  <resprocessing>
    <respcondition><conditionvar><varequal respident="RESP_1">OPT_A</varequal></conditionvar><setvar varname="D2L_Correct" action="Add">1</setvar></respcondition>
    <respcondition><conditionvar><varequal respident="RESP_1">OPT_B</varequal></conditionvar><setvar varname="D2L_Incorrect" action="Add">1</setvar></respcondition>
    <respcondition><conditionvar><varequal respident="RESP_2">OPT_A</varequal></conditionvar><setvar varname="D2L_Incorrect" action="Add">1</setvar></respcondition>
    <respcondition><conditionvar><varequal respident="RESP_2">OPT_B</varequal></conditionvar><setvar varname="D2L_Correct" action="Add">1</setvar></respcondition>
  </resprocessing>
</item>
"""
    )
    row, diagnostics = extractor.parse_item(item, "inline", "quiz.xml", "", "assessment/section[S1]/item")
    pairs = json.loads(row["matching_pairs"])

    assert row["response_schema"] == "matching"
    assert row["stem_text"] == "Match the pairs."
    assert pairs == [
        {"prompt": "First prompt", "correct": ["Alpha"], "respident": "RESP_1"},
        {"prompt": "Second prompt", "correct": ["Beta"], "respident": "RESP_2"},
    ]
    extractor.enrich_question_rows_for_review([row])
    assert row["matching_review_display"] == "First prompt -> Alpha\nSecond prompt -> Beta"
    assert not diagnostics


def test_parse_short_answer_multiple_answers_and_case_flag():
    item = item_from_xml(
        """
<item label="SHORT_1" title="Plural of appendix">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Short Answer</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_weighting</fieldlabel><fieldentry>1.000000000</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation>
    <flow>
      <material><mattext texttype="text/html"><p>What is the plural of appendix?</p></mattext></material>
      <response_str ident="SHORT_1_STR" rcardinality="Single">
        <render_fib rows="1" columns="20" prompt="Box" fibtype="String">
          <response_label ident="SHORT_1_ANS" />
        </render_fib>
      </response_str>
    </flow>
  </presentation>
  <resprocessing>
    <respcondition><conditionvar><varequal respident="SHORT_1_ANS" case="no">appendices</varequal></conditionvar><setvar action="Set">100.000000000</setvar></respcondition>
    <respcondition><conditionvar><varequal respident="SHORT_1_ANS" case="no">Appendices</varequal></conditionvar><setvar action="Set">100.000000000</setvar></respcondition>
  </resprocessing>
</item>
"""
    )
    row, diagnostics = extractor.parse_item(item, "inline", "quiz.xml", "", "assessment/section[S1]/item")
    answers = json.loads(row["accepted_answers"])

    assert row["response_schema"] == "short_text"
    assert row["correct_answer_text"] == "BLANK_1: appendices, Appendices"
    assert answers[0]["accepted_values"] == ["appendices", "Appendices"]
    assert answers[0]["case_sensitive"] == "no"
    assert not diagnostics


def test_parse_fill_in_blank_with_interleaved_material():
    item = item_from_xml(
        """
<item label="FITB_1">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Fill in the Blanks</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_weighting</fieldlabel><fieldentry>1.000000000</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation>
    <flow>
      <material><mattext texttype="text/html">Blood is </mattext></material>
      <response_str ident="FITB_1_STR" rcardinality="Single">
        <render_fib rows="1" columns="15" prompt="Box" fibtype="String">
          <response_label ident="FITB_1_ANS" />
        </render_fib>
      </response_str>
      <material><mattext texttype="text/html"> in color.</mattext></material>
    </flow>
  </presentation>
  <resprocessing>
    <respcondition><conditionvar><varequal respident="FITB_1_ANS" case="no">red</varequal></conditionvar><setvar action="Set">100.000000000</setvar></respcondition>
  </resprocessing>
</item>
"""
    )
    row, diagnostics = extractor.parse_item(item, "inline", "quiz.xml", "", "assessment/section[S1]/item")

    assert row["response_schema"] == "fill_in_blank"
    assert row["stem_text"] == "Blood is [BLANK_1] in color."
    assert row["correct_answer_text"] == "BLANK_1: red"
    assert not diagnostics


def test_parse_numeric_tolerance_best_effort():
    item = item_from_xml(
        """
<item label="NUM_1" title="Numeric Question">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Numeric</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_weighting</fieldlabel><fieldentry>2.000000000</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation>
    <flow>
      <material><mattext texttype="text/html">Enter a number.</mattext></material>
      <response_num ident="NUM_1_NUM" rcardinality="Single" />
    </flow>
  </presentation>
  <resprocessing>
    <respcondition><conditionvar><vargte respident="NUM_1_NUM">9.5</vargte><varlte respident="NUM_1_NUM">10.5</varlte></conditionvar><setvar action="Set">100.000000000</setvar></respcondition>
  </resprocessing>
</item>
"""
    )
    row, diagnostics = extractor.parse_item(item, "inline", "quiz.xml", "", "assessment/section[S1]/item")
    tolerance = json.loads(row["numeric_tolerance"])

    assert row["response_schema"] == "numeric"
    assert row["numeric_answer"] == ""
    assert {entry["operator"] for entry in tolerance} == {"vargte", "varlte"}
    assert any(diag["issue_type"] == "numeric_best_effort" for diag in diagnostics)


def test_parse_ordering_uses_best_effort_when_sequence_is_missing():
    item = item_from_xml(
        """
<item label="ORD_1" title="Ordering Question">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Ordering</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_weighting</fieldlabel><fieldentry>1.000000000</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation>
    <flow>
      <material><mattext texttype="text/html">Place the items in order.</mattext></material>
      <response_lid ident="ORD_1_LID" rcardinality="Multiple">
        <render_choice shuffle="no">
          <flow_label class="Block">
            <response_label ident="O_A"><flow_mat><material><mattext texttype="text/html">First</mattext></material></flow_mat></response_label>
            <response_label ident="O_B"><flow_mat><material><mattext texttype="text/html">Second</mattext></material></flow_mat></response_label>
            <response_label ident="O_C"><flow_mat><material><mattext texttype="text/html">Third</mattext></material></flow_mat></response_label>
          </flow_label>
        </render_choice>
      </response_lid>
    </flow>
  </presentation>
  <resprocessing>
    <respcondition><conditionvar><varequal respident="ORD_1_LID">O_A</varequal></conditionvar><setvar action="Set">0.000000000</setvar></respcondition>
  </resprocessing>
</item>
"""
    )
    row, diagnostics = extractor.parse_item(item, "inline", "quiz.xml", "", "assessment/section[S1]/item")
    extractor.enrich_question_rows_for_review([row])

    assert row["ordering_sequence"] == "A;B;C"
    assert row["ordering_review_display"].startswith("Best-effort sequence")
    assert any(diag["issue_type"] == "ordering_best_effort" for diag in diagnostics)


def test_parse_ordering_response_group_reconstructs_sequence_from_positions():
    item = item_from_xml(
        """
<item label="ORD_GRP_1" title="QUES_9000">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Ordering</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_weighting</fieldlabel><fieldentry>1.000000000</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation>
    <flow>
      <material><mattext texttype="text/html">Put the steps in order.</mattext></material>
      <response_grp respident="ORD_GRP" rcardinality="Ordered">
        <render_choice shuffle="yes">
          <flow_label class="Block">
            <response_label ident="STEP_A"><flow_mat><material><mattext texttype="text/html">First step</mattext></material></flow_mat></response_label>
            <response_label ident="STEP_B"><flow_mat><material><mattext texttype="text/html">Second step</mattext></material></flow_mat></response_label>
            <response_label ident="STEP_C"><flow_mat><material><mattext texttype="text/html">Third step</mattext></material></flow_mat></response_label>
          </flow_label>
        </render_choice>
      </response_grp>
    </flow>
  </presentation>
  <resprocessing>
    <respcondition><conditionvar><varequal respident="STEP_A">2</varequal></conditionvar><setvar varname="D2L_Correct" action="Add">1</setvar></respcondition>
    <respcondition><conditionvar><varequal respident="STEP_B">3</varequal></conditionvar><setvar varname="D2L_Correct" action="Add">1</setvar></respcondition>
    <respcondition><conditionvar><varequal respident="STEP_C">1</varequal></conditionvar><setvar varname="D2L_Correct" action="Add">1</setvar></respcondition>
  </resprocessing>
</item>
"""
    )

    row, diagnostics = extractor.parse_item(item, "inline", "quiz.xml", "", "assessment/section[S1]/item")
    extractor.enrich_question_rows_for_review([row])

    assert row["ordering_sequence"] == "C;A;B"
    assert row["correct_answer_text"] == "Third step | First step | Second step"
    assert row["ordering_review_display"] == "1. Third step\n2. First step\n3. Second step"
    assert not diagnostics


def test_feedback_channels_remain_separate_when_possible():
    item = item_from_xml(
        """
<item label="MC_1" title="Feedback Question">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Multiple Choice</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_weighting</fieldlabel><fieldentry>1.000000000</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation>
    <flow>
      <material><mattext texttype="text/html">Pick the correct answer.</mattext></material>
      <response_lid ident="MC_1_LID" rcardinality="Single">
        <render_choice shuffle="no">
          <flow_label class="Block">
            <response_label ident="ANS_A"><flow_mat><material><mattext texttype="text/html">Alpha</mattext></material></flow_mat></response_label>
            <response_label ident="ANS_B"><flow_mat><material><mattext texttype="text/html">Beta</mattext></material></flow_mat></response_label>
          </flow_label>
        </render_choice>
      </response_lid>
    </flow>
  </presentation>
  <resprocessing>
    <respcondition><conditionvar><varequal respident="MC_1_LID">ANS_A</varequal></conditionvar><setvar action="Set">100.000000000</setvar><displayfeedback feedbacktype="Response" linkrefid="fb_correct" /></respcondition>
    <respcondition><conditionvar><varequal respident="MC_1_LID">ANS_B</varequal></conditionvar><setvar action="Set">0.000000000</setvar><displayfeedback feedbacktype="Response" linkrefid="fb_incorrect" /></respcondition>
  </resprocessing>
  <itemfeedback ident="fb_correct"><material><mattext texttype="text/plain">Correct answer.</mattext></material></itemfeedback>
  <itemfeedback ident="fb_incorrect"><material><mattext texttype="text/plain">Try again.</mattext></material></itemfeedback>
  <itemfeedback ident="fb_general"><material><mattext texttype="text/plain">Review chapter 1.</mattext></material></itemfeedback>
</item>
"""
    )
    row, diagnostics = extractor.parse_item(item, "inline", "quiz.xml", "", "assessment/section[S1]/item")

    assert row["correct_feedback"] == "Correct answer."
    assert row["incorrect_feedback"] == "Try again."
    assert row["general_feedback"] == "Review chapter 1."
    assert "A: Correct answer." in row["answer_specific_feedback"]
    assert "B: Try again." in row["answer_specific_feedback"]
    assert any(diag["issue_type"] == "feedback_partial_collapse" for diag in diagnostics)


def test_question_title_review_fallback_only_for_id_like_titles():
    assert extractor.derive_question_title_review("Short Answer Prompt", "QUES_1", "Ignored stem") == "Short Answer Prompt"
    assert extractor.derive_question_title_review("QUES_1", "QUES_1", "Explain the process in detail. [BLANK_1]") == "Explain the process in detail."
    assert extractor.derive_question_title_review("ITEM_999", "QUES_2", "Describe osmosis across the membrane.") == "Describe osmosis across the membrane."
    assert extractor.derive_question_title_review("", "QUES_3", "") == "QUES_3"


def test_ordering_review_display_uses_placeholder_labels_when_text_is_missing():
    row = {
        "question_type": "Ordering",
        "ordering_sequence": "A;B",
        "question_payload_json": json.dumps(
            {
                "choices": {
                    "A": {"ident": "STEP_A", "text": ""},
                    "B": {"ident": "STEP_B", "text": ""},
                },
                "sequence": ["A", "B"],
                "best_effort": False,
                "conditions": [],
            }
        ),
    }

    extractor.enrich_question_rows_for_review([row])

    assert row["ordering_review_display"] == "1. Item 1\n2. Item 2"


def test_question_image_refs_are_aggregated_and_resolved_across_blocks(tmp_path: Path):
    (tmp_path / "csfiles" / "home_dir").mkdir(parents=True)
    (tmp_path / "csfiles" / "home_dir" / "stem.png").write_text("stem", encoding="utf-8")
    (tmp_path / "csfiles" / "home_dir" / "choice.png").write_text("choice", encoding="utf-8")
    (tmp_path / "root.png").write_text("root", encoding="utf-8")

    item = item_from_xml(
        """
<item label="IMG_1" title="Image Question">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Multiple Choice</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_weighting</fieldlabel><fieldentry>1.000000000</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation>
    <flow>
      <material><mattext texttype="text/html">&lt;p&gt;Stem &lt;img src="csfiles\\home_dir\\stem.png?download=1" /&gt;&lt;/p&gt;</mattext></material>
      <response_lid ident="IMG_1_LID" rcardinality="Single">
        <render_choice shuffle="no">
          <flow_label class="Block">
            <response_label ident="A1">
              <flow_mat><material><mattext texttype="text/html">&lt;img src="https://example.edu/content/view?item=1#/csfiles/home_dir/choice.png#frag" /&gt;Choice A</mattext></material></flow_mat>
            </response_label>
            <response_label ident="A2">
              <flow_mat><material><mattext texttype="text/html">Choice B</mattext></material></flow_mat>
            </response_label>
          </flow_label>
        </render_choice>
      </response_lid>
    </flow>
  </presentation>
  <resprocessing>
    <respcondition><conditionvar><varequal respident="IMG_1_LID">A1</varequal></conditionvar><setvar action="Set">100.000000000</setvar></respcondition>
  </resprocessing>
  <itemfeedback ident="fb_1"><material><mattext texttype="text/html">&lt;p&gt;&lt;img src="missing.png" /&gt;&lt;/p&gt;</mattext></material></itemfeedback>
  <answer_key><answer_key_material><flow_mat><material><matimage uri="root.png" /></material></flow_mat></answer_key_material></answer_key>
</item>
"""
    )

    row, diagnostics = extractor.parse_item(item, "inline", "quiz.xml", "", "assessment/section[SECT_1]/item")
    diagnostics.extend(
        extractor.populate_row_image_fields(
            row,
            file_index=extractor.build_export_file_index(tmp_path),
            source_file="quiz.xml",
        )
    )

    assert row["image_refs"].split(";") == [
        r"csfiles\home_dir\stem.png?download=1",
        "https://example.edu/content/view?item=1#/csfiles/home_dir/choice.png#frag",
        "missing.png",
        "root.png",
    ]
    assert row["image_paths_resolved"].split(";") == [
        "csfiles/home_dir/stem.png",
        "csfiles/home_dir/choice.png",
        "root.png",
    ]
    assert row["image_count"] == 3
    issue_types = [diag["issue_type"] for diag in diagnostics]
    assert "missing_image_file" in issue_types
    assert "unresolved_image_ref" in issue_types


def test_resolve_image_ref_prefers_deterministic_order_and_reports_duplicates(tmp_path: Path):
    (tmp_path / "csfiles" / "home_dir").mkdir(parents=True)
    (tmp_path / "csfiles" / "home_dir" / "priority.png").write_text("home", encoding="utf-8")
    (tmp_path / "priority.png").write_text("root", encoding="utf-8")
    (tmp_path / "root-only.png").write_text("root", encoding="utf-8")
    (tmp_path / "folder").mkdir()
    (tmp_path / "folder" / "fallback.png").write_text("fallback", encoding="utf-8")
    (tmp_path / "dup_a").mkdir()
    (tmp_path / "dup_b").mkdir()
    (tmp_path / "dup_a" / "duplicate.png").write_text("a", encoding="utf-8")
    (tmp_path / "dup_b" / "duplicate.png").write_text("b", encoding="utf-8")

    file_index = extractor.build_export_file_index(tmp_path)

    resolved_priority, priority_diags = extractor.resolve_image_ref(
        "priority.png",
        file_index,
        question_id="Q1",
        source_file="quiz.xml",
        source_hint="assessment/section[SECT_1]/item",
    )
    resolved_root, root_diags = extractor.resolve_image_ref(
        "root-only.png",
        file_index,
        question_id="Q1",
        source_file="quiz.xml",
        source_hint="assessment/section[SECT_1]/item",
    )
    resolved_fallback, fallback_diags = extractor.resolve_image_ref(
        "fallback.png",
        file_index,
        question_id="Q1",
        source_file="quiz.xml",
        source_hint="assessment/section[SECT_1]/item",
    )
    unresolved_duplicate, duplicate_diags = extractor.resolve_image_ref(
        "duplicate.png",
        file_index,
        question_id="Q1",
        source_file="quiz.xml",
        source_hint="assessment/section[SECT_1]/item",
    )

    assert resolved_priority == "csfiles/home_dir/priority.png"
    assert not priority_diags
    assert resolved_root == "root-only.png"
    assert not root_diags
    assert resolved_fallback == "folder/fallback.png"
    assert not fallback_diags
    assert unresolved_duplicate == ""
    assert {diag["issue_type"] for diag in duplicate_diags} == {"duplicate_image_filename", "unresolved_image_ref"}


def test_question_bank_matching_supports_exact_ambiguous_and_none(tmp_path: Path):
    write_questiondb(
        tmp_path,
        """
<section ident="SECT_WEEK1" title="rsp_MEDT_1000_Week_1_Quiz_rb1">
  <item label="BANK_EXACT" title="Bank Exact">
    <itemmetadata>
      <qtimetadata>
        <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Multiple Choice</fieldentry></qti_metadatafield>
        <qti_metadatafield><fieldlabel>qmd_displayid</fieldlabel><fieldentry>DISPLAY-1</fieldentry></qti_metadatafield>
      </qtimetadata>
    </itemmetadata>
    <presentation><flow><material><mattext texttype="text/html">Exact stem</mattext></material></flow></presentation>
  </item>
  <item label="BANK_AMBIG_1" title="Repeated Title">
    <itemmetadata><qtimetadata><qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Short Answer</fieldentry></qti_metadatafield></qtimetadata></itemmetadata>
    <presentation><flow><material><mattext texttype="text/html">Repeated stem</mattext></material></flow></presentation>
  </item>
</section>
<section ident="SECT_WEEK2" title="rsp_MEDT_1000_Week_2_Quiz_rb1">
  <item label="BANK_AMBIG_2" title="Repeated Title">
    <itemmetadata><qtimetadata><qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Short Answer</fieldentry></qti_metadatafield></qtimetadata></itemmetadata>
    <presentation><flow><material><mattext texttype="text/html">Repeated stem</mattext></material></flow></presentation>
  </item>
</section>
""",
    )
    bank, _ = extractor.parse_questiondb(tmp_path)

    exact_item = item_from_xml(
        """
<item label="INLINE_EXACT" title="Inline Exact">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Multiple Choice</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_displayid</fieldlabel><fieldentry>DISPLAY-1</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation><flow><material><mattext texttype="text/html">Exact stem</mattext></material></flow></presentation>
</item>
"""
    )
    exact_row, _ = extractor.parse_item(exact_item, "inline", "quiz.xml", "", "assessment/section[S1]/item")
    exact_match = bank.match_inline_item(exact_item, exact_row, "Week 1 Quiz")

    ambiguous_item = item_from_xml(
        """
<item label="INLINE_AMBIG" title="Repeated Title">
  <itemmetadata><qtimetadata><qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Short Answer</fieldentry></qti_metadatafield></qtimetadata></itemmetadata>
  <presentation><flow><material><mattext texttype="text/html">Repeated stem</mattext></material></flow></presentation>
</item>
"""
    )
    ambiguous_row, _ = extractor.parse_item(ambiguous_item, "inline", "quiz.xml", "", "assessment/section[S1]/item")
    ambiguous_match = bank.match_inline_item(ambiguous_item, ambiguous_row, "Practice Quiz")

    none_item = item_from_xml(
        """
<item label="INLINE_NONE" title="Different Title">
  <itemmetadata><qtimetadata><qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Short Answer</fieldentry></qti_metadatafield></qtimetadata></itemmetadata>
  <presentation><flow><material><mattext texttype="text/html">Different stem</mattext></material></flow></presentation>
</item>
"""
    )
    none_row, _ = extractor.parse_item(none_item, "inline", "quiz.xml", "", "assessment/section[S1]/item")
    none_match = bank.match_inline_item(none_item, none_row, "Practice Quiz")

    assert exact_match["status"] == "matched"
    assert exact_match["reason"] == "exact stable key match"
    assert ambiguous_match["status"] == "ambiguous"
    assert none_match["status"] == "none"


def test_create_review_outputs_adds_image_columns_and_hyperlinks(tmp_path: Path):
    export_root = create_minimal_export(
        tmp_path,
        """
<item label="IMG_OUT_1" title="Image Output Question">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Multiple Choice</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_weighting</fieldlabel><fieldentry>1.000000000</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation>
    <flow>
      <material><mattext texttype="text/html">&lt;p&gt;Identify the structure.&lt;img src="picture.png" /&gt;&lt;/p&gt;</mattext></material>
      <response_lid ident="IMG_OUT_LID" rcardinality="Single">
        <render_choice shuffle="no">
          <flow_label class="Block">
            <response_label ident="A1"><flow_mat><material><mattext texttype="text/html">Alpha</mattext></material></flow_mat></response_label>
            <response_label ident="A2"><flow_mat><material><mattext texttype="text/html">Beta</mattext></material></flow_mat></response_label>
          </flow_label>
        </render_choice>
      </response_lid>
    </flow>
  </presentation>
  <resprocessing>
    <respcondition><conditionvar><varequal respident="IMG_OUT_LID">A1</varequal></conditionvar><setvar action="Set">100.000000000</setvar></respcondition>
  </resprocessing>
</item>
""",
    )
    (export_root / "csfiles" / "home_dir").mkdir(parents=True)
    (export_root / "csfiles" / "home_dir" / "picture.png").write_text("img", encoding="utf-8")

    output_root = tmp_path / "review"
    xlsx_path = extractor.create_review_outputs(export_root, output_root)
    data = json.loads((output_root / "quiz_review.json").read_text(encoding="utf-8"))
    question = data["questions"][0]

    assert question["image_refs"] == "picture.png"
    assert question["image_paths_resolved"] == "csfiles/home_dir/picture.png"
    assert question["image_count"] == 1
    assert question["image_link_primary"].endswith("csfiles/home_dir/picture.png")

    workbook = load_workbook(xlsx_path)
    worksheet = workbook["questions"]
    headers = [cell.value for cell in worksheet[1]]
    assert headers[:12] == [
        "quiz_id",
        "quiz_title",
        "section_id",
        "section_title",
        "question_order",
        "question_id",
        "question_title_review",
        "question_type",
        "points",
        "has_image",
        "image_count",
        "image_link_primary",
    ]
    assert headers[headers.index("question_title") + 1 : headers.index("question_title") + 5] == [
        "source_location",
        "source_hint",
        "source_quiz_file",
        "source_bank_file",
    ]
    assert question["has_image"] == "yes"
    assert question["question_title_review"] == "Image Output Question"
    image_link_cell = worksheet.cell(row=2, column=headers.index("image_link_primary") + 1)
    assert image_link_cell.hyperlink is not None
    assert image_link_cell.hyperlink.target.endswith("csfiles/home_dir/picture.png")


def test_create_review_outputs_adds_matching_pairs_expanded_sheet_and_json(tmp_path: Path):
    export_root = create_minimal_export(
        tmp_path,
        """
<item label="MATCH_OUT_1" title="QUES_1200">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Matching</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_weighting</fieldlabel><fieldentry>2.000000000</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation>
    <flow>
      <material><mattext texttype="text/html">Match the labels.</mattext></material>
      <response_grp respident="RESP_1" rcardinality="Single">
        <material><mattext texttype="text/html">A</mattext></material>
        <render_choice shuffle="yes">
          <flow_label class="Block">
            <response_label ident="OPT_1"><flow_mat><material><mattext texttype="text/html">Alpha</mattext></material></flow_mat></response_label>
            <response_label ident="OPT_2"><flow_mat><material><mattext texttype="text/html">Beta</mattext></material></flow_mat></response_label>
          </flow_label>
        </render_choice>
      </response_grp>
      <response_grp respident="RESP_2" rcardinality="Single">
        <material><mattext texttype="text/html">B</mattext></material>
        <render_choice shuffle="yes">
          <flow_label class="Block">
            <response_label ident="OPT_1"><flow_mat><material><mattext texttype="text/html">Alpha</mattext></material></flow_mat></response_label>
            <response_label ident="OPT_2"><flow_mat><material><mattext texttype="text/html">Beta</mattext></material></flow_mat></response_label>
          </flow_label>
        </render_choice>
      </response_grp>
    </flow>
  </presentation>
  <resprocessing>
    <respcondition><conditionvar><varequal respident="RESP_1">OPT_2</varequal></conditionvar><setvar varname="D2L_Correct" action="Add">1</setvar></respcondition>
    <respcondition><conditionvar><varequal respident="RESP_2">OPT_1</varequal></conditionvar><setvar varname="D2L_Correct" action="Add">1</setvar></respcondition>
  </resprocessing>
</item>
""",
    )

    output_root = tmp_path / "review_matching"
    xlsx_path = extractor.create_review_outputs(export_root, output_root)
    data = json.loads((output_root / "quiz_review.json").read_text(encoding="utf-8"))

    expanded_rows = data["matching_pairs_expanded"]
    assert len(expanded_rows) == 2
    assert expanded_rows[0]["prompt_order"] == 1
    assert expanded_rows[0]["correct_match"] == "Beta"
    assert expanded_rows[1]["prompt_order"] == 2
    assert expanded_rows[1]["correct_match"] == "Alpha"
    assert expanded_rows[0]["question_title_review"] == "Match the labels."

    workbook = load_workbook(xlsx_path)
    assert "matching_pairs_expanded" in workbook.sheetnames
    worksheet = workbook["matching_pairs_expanded"]
    headers = [cell.value for cell in worksheet[1]]
    assert headers[:12] == [
        "quiz_id",
        "quiz_title",
        "section_id",
        "section_title",
        "question_order",
        "question_id",
        "question_title_review",
        "question_title",
        "prompt_order",
        "match_order",
        "prompt",
        "correct_match",
    ]


def test_create_review_outputs_copy_images_to_assets_copies_and_relinks(tmp_path: Path):
    export_root = create_minimal_export(
        tmp_path,
        """
<item label="IMG_COPY_1" title="Image Copy Question">
  <itemmetadata>
    <qtimetadata>
      <qti_metadatafield><fieldlabel>qmd_questiontype</fieldlabel><fieldentry>Multiple Choice</fieldentry></qti_metadatafield>
      <qti_metadatafield><fieldlabel>qmd_weighting</fieldlabel><fieldentry>1.000000000</fieldentry></qti_metadatafield>
    </qtimetadata>
  </itemmetadata>
  <presentation>
    <flow>
      <material><mattext texttype="text/html">&lt;p&gt;Question.&lt;img src="copy_me.png" /&gt;&lt;/p&gt;</mattext></material>
      <response_lid ident="IMG_COPY_LID" rcardinality="Single">
        <render_choice shuffle="no">
          <flow_label class="Block">
            <response_label ident="A1"><flow_mat><material><mattext texttype="text/html">Yes</mattext></material></flow_mat></response_label>
            <response_label ident="A2"><flow_mat><material><mattext texttype="text/html">No</mattext></material></flow_mat></response_label>
          </flow_label>
        </render_choice>
      </response_lid>
    </flow>
  </presentation>
  <resprocessing>
    <respcondition><conditionvar><varequal respident="IMG_COPY_LID">A1</varequal></conditionvar><setvar action="Set">100.000000000</setvar></respcondition>
  </resprocessing>
</item>
""",
    )
    (export_root / "csfiles" / "home_dir").mkdir(parents=True)
    (export_root / "csfiles" / "home_dir" / "copy_me.png").write_text("img", encoding="utf-8")

    output_root = tmp_path / "review_copy"
    xlsx_path = extractor.create_review_outputs(export_root, output_root, copy_images_to_assets=True)
    data = json.loads((output_root / "quiz_review.json").read_text(encoding="utf-8"))
    question = data["questions"][0]
    copied_file = output_root / "assets" / "csfiles" / "home_dir" / "copy_me.png"

    assert copied_file.exists()
    assert question["image_paths_resolved"] == "csfiles/home_dir/copy_me.png"
    assert question["image_link_primary"] == "assets/csfiles/home_dir/copy_me.png"

    workbook = load_workbook(xlsx_path)
    worksheet = workbook["questions"]
    headers = [cell.value for cell in worksheet[1]]
    image_link_cell = worksheet.cell(row=2, column=headers.index("image_link_primary") + 1)
    assert image_link_cell.hyperlink is not None
    assert image_link_cell.hyperlink.target == "assets/csfiles/home_dir/copy_me.png"


def test_integration_sample_export_outputs_have_traceability_and_filters(tmp_path: Path):
    sample_root = REPO_ROOT / "Sample_extracted_export_files" / "MEDT 1000"
    if not sample_root.exists():
        sample_exports_root = REPO_ROOT / "Sample_extracted_export_files"
        if not sample_exports_root.exists():
            pytest.skip("Bundled sample export fixture is not included in this source checkout.")
        sample_root = next(
            (
                manifest.parent
                for manifest in sample_exports_root.rglob("imsmanifest.xml")
                if list(manifest.parent.glob("quiz_d2l_*.xml"))
            ),
            None,
        )
        if sample_root is None:
            pytest.skip("Bundled sample export fixture is not available.")
    output_root = tmp_path / "generated_review"

    xlsx_path = extractor.create_review_outputs(sample_root, output_root)
    data = json.loads((output_root / "quiz_review.json").read_text(encoding="utf-8"))

    assert all(row["quiz_title"] for row in data["sections_pools"])
    assert all(row["quiz_title"] for row in data["questions"])
    assert all(row["quiz_title"] for row in data["pool_members"])
    assert all(row["quiz_title"] for row in data["diagnostics"])
    assert all("source_hint" in row for row in data["source_map"])
    assert "source_xpath_or_hint" not in data["source_map"][0]
    assert any(row["storage_type"] == "inline" for row in data["quiz_overview"])
    assert any(row["storage_type"] == "hybrid" for row in data["quiz_overview"])
    assert any(row["source_location"] == "hybrid" for row in data["questions"])
    assert any(row["source_location"] == "inline" for row in data["questions"])
    assert all(row["source_location"] != "inline_quiz" for row in data["questions"])
    assert all("image_refs" in row for row in data["questions"])
    assert all("image_paths_resolved" in row for row in data["questions"])
    assert all("image_count" in row for row in data["questions"])
    assert all("image_link_primary" in row for row in data["questions"])
    assert all("question_title_review" in row for row in data["questions"])
    assert all("matching_review_display" in row for row in data["questions"])
    assert all("ordering_review_display" in row for row in data["questions"])
    assert all("has_image" in row for row in data["questions"])
    assert "matching_pairs_expanded" in data
    assert any(row["draw_count"] not in ("", None) for row in data["sections_pools"] if row["section_type"] == "pool")
    assert any(row["pool_size"] not in ("", None) for row in data["pool_members"])
    assert any(row["issue_type"] == "pool_dependency" for row in data["diagnostics"])

    workbook = load_workbook(xlsx_path)
    assert "matching_pairs_expanded" in workbook.sheetnames
    for worksheet in workbook.worksheets:
        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref
